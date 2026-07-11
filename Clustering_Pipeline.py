"""
Codon-usage clustering pipeline (FASTA input + optional automatic/basic cluster extraction
+ optional DAVID sliding-window enrichment scan).
"""

import os
import re
import sys
import shutil
import time
import gc
import threading
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
try:
    from tkinter import Tk, messagebox
    from tkinter.filedialog import askopenfilename
except Exception:
    Tk = None
    messagebox = None
    askopenfilename = None

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = HERE
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from codonpipe.clustering import run_codon_clustering, choose_cluster_file, render_trna_gene_ordered_heatmaps
from codonpipe.fasta_metrics import (
    auto_find_fasta, choose_fasta, build_locus_index,
    canonicalize_id, enforce_unique_after_canon,
    canonicalize_cluster_df, canonicalize_generic_map,
    build_metrics_table, build_summary, build_quantitative_bis,
    build_fasta_metric_cluster_df, append_fasta_metric_clusters
)
from codonpipe.excel_outputs import (
    build_coordinates_df,
    build_locus_tags_sheet, build_binary_sheet,
    build_meta_table, write_per_cluster_workbook,
    write_codon_usage_by_cluster_workbook,
    write_trna_abundance_correlation_outputs,
)
from codonpipe.ks2d import compute_2d_ks_for_clusters
from codonpipe.density_bridge import run_density_plot_script
from codonpipe.legend import write_all_texts
from codonpipe.david_window_scan import run_david_window_scan_from_ordered_genes


# =============================================================================
# USER SETTINGS
# =============================================================================

SET = dict(
    # Paths / dataset context
    default_root="",

    # FASTA input
    fasta_path='',
    fasta_row_id_mode='primary',         # 'primary' | 'old' | 'locus'
    fasta_trim_to_multiple_of_3=True,
    fasta_include_stops=None,            # None -> infer from codon_set (64_all keeps stops)
    fasta_codon_table_mode='abs',        # harmonized clustering/export path uses raw codon counts
    fasta_codon_range='all',             # 'all' or 1-based inclusive codon range, e.g. '1-20', '20-200', '20-end'
    fasta_write_intermediate_excels=True,

    # Optional FASTA-derived metric clusters appended to the selected cluster table
    fasta_metric_clusters_enable=False,
    fasta_metric_cluster_configs=[],
    fasta_metric_cluster_scores_path='',
    fasta_metric_cluster_file_path='',

    # Optional custom CDS appended to the selected genome before analysis
    custom_cds_enable=False,
    custom_cds_paths=[],                 # list of extra FASTA/.fna files, each containing >=1 CDS
    custom_cds_include_as_cluster=False,
    custom_cds_cluster_name='custom',
    custom_cds_generated_ids=[],
    custom_cds_original_fasta_path='',
    custom_cds_merged_fasta_path='',

    # UMAP parameters
    umap_neighbors=10,
    umap_min_dist=0.01,
    umap_metric='cosine',
    umap_components=2,
    umap_randomize=False,
    umap_clip_abs=0.0,
    umap_init='spectral',
    fasta_organism_mode='prokaryote',

    # t-SNE parameters
    tsne_perplexity=10,
    tsne_distance='cosine',
    tsne_dims=2,
    tsne_exaggeration=10,
    tsne_learnrate=100,

    # Scatter plot parameters
    scatter_fig_size=(4, 4),
    scatter_point_size=2.5,
    scatter_point_alpha=0.7,
    scatter_edge_width=0.1,
    scatter_color_mode='enrichment',

    # Density/enrichment params
    density_nbins=150,
    density_sigma=4.0,
    density_use_log=True,
    density_min_rel=0.00,
    density_max_rel=1.00,
    density_cmap_name='plasma_r',
    density_enrichment_eps=1e-12,
    density_enrichment_use_log=True,

    # Automatic/basic cluster extraction + plot-display filters
    plot_cluster_min_genes=20,
    basic_cluster_output_id_column='RefSeq_LocusTag_RS',
    basic_cluster_output_id_fallbacks=['GeneSymbol', 'LocusTag', 'PrimaryID'],
    basic_cluster_keyword_groups=None,  # None -> use BASIC_CLUSTER_KEYWORD_GROUPS; GUI can provide an edited dict
    basic_cluster_search_columns=[
        'GeneSymbol', 'ProteinDescription', 'UniProtID',
        'RefSeqProteinID', 'RefSeq_LocusTag_RS', 'Old_LocusTag'
    ],

    # DAVID sliding-window scan
    david_user_email='',  # optional DAVID-registered email; leave empty to be asked/interactively supplied if a DAVID scan is requested
    david_window_size=100,
    david_step_size=50,
    david_wait_time=0.0,
    david_max_clusters=3,
    david_min_valid_ids_per_window=3,
    david_plot_format='png',
    david_report_subdir_name='DAVID window reports',
    david_top_n_hits=10,  # top N sliding-window hits used both for output and whole-genome term reconstruction

    # DAVID whole-genome term reconstruction
    david_chart_threshold=1.0,
    david_chart_count=1,
    david_manual_term_queries=[],              # leave empty to use auto queries from top sliding-window hits
    david_term_match_mode='contains',          # 'contains' or 'exact'
    david_append_terms_to_geneids_excel=True,
    david_gene2terms_filename='DAVID gene2terms.txt',
    david_gene2terms_path='',             # optional explicit path selected from GUI

    # Gene-cluster localization heatmap settings
    gchm_enable=True,
    gchm_colormap='plasma',
    gchm_custom_cmaps_xlsx="",
    gchm_sigma=10,
    gchm_spread_factor=5,
    gchm_height_per_cluster=0.3,
    gchm_label_fontsize=10,
    gchm_dpi=300,
    gchm_cmap_min_rel=0.2,
    gchm_cmap_max_rel=1.0,
    gchm_output_filename='gene_cluster_heatmap_KS.png',
    gchm_show_fig=True,

    # Optional figure axis limits (blank / None keeps automatic limits)
    heatmap_xmin=None,
    heatmap_xmax=None,
    heatmap_ymin=None,
    heatmap_ymax=None,
    scatter_xmin=None,
    scatter_xmax=None,
    scatter_ymin=None,
    scatter_ymax=None,
    trna_abundance_heatmap_xmin=None,
    trna_abundance_heatmap_xmax=None,
    trna_abundance_heatmap_ymin=None,
    trna_abundance_heatmap_ymax=None,
    trna_abundance_scatter_xmin=None,
    trna_abundance_scatter_xmax=None,
    trna_abundance_scatter_ymin=None,
    trna_abundance_scatter_ymax=None,

    # PCA parameters
    pca_npcs=3,
    pca_center=True,
    pca_scale=True,

    # Gene ordering / clustering along genome
    gene_dist_metric='euclidean',
    gene_linkage='single',
    cluster_use_optimal_leaf_ordering=True,
    cluster_optimal_leaf_max_size=250,
    cluster_fast_order_threshold=800,
    hierarchical_optimal_leaf_max_size=400,
    hierarchical_fast_order_threshold=2000,

    # Cluster method parameters
    kmeans_k=12,
    kmedoids_k=12,
    kmedoids_dist='euclidean',
    spectral_k=12,
    spectral_dist='euclidean',
    dbscan_eps=1.5,
    dbscan_minpts=10,
    dbscan_dist='euclidean',

    # Feature normalization
    center_features=True,
    scale_features=True,

    # Feature clustering
    feature_dist_metric='spearman',
    feature_linkage='single',

    # Smoothing / binning along genome axis
    apply_smoothing=True,
    smooth_window_genes=6,
    apply_binning=False,
    bin_size_genes=50,

    # Heatmap aesthetics
    heatmap_colormap_name='parula',
    heatmap_caxis_limits=(-0.5, 2.5),
    heatmap_fig_size=(18, 4),
    xtick_every_genes=500,

    # Typography / export
    font_name='Arial',
    font_size_axes=7,
    font_size_xticks=8,
    font_size_yticks=3,
    font_size_titles=10,
    colorbar_title_size=11,
    figure_dpi=300,

    # Titles / labels
    heatmap_title_template='Heatmap — {USAGE} | {CSET} | {DIMRED} | {CLUSTER}',
    scatter_title_template='{DIMRED} scatter — {USAGE} | {CSET}',
    colorbar_title_string=r'$\sigma_{\mathrm{codon}}$',

    # Save formats
    save_png=True,
    save_pdf=False,
    save_jpeg=False,

    # Custom colormaps (optional)
    use_custom_colormaps=False,
    custom_cmap_excel="",
    custom_cmap_sheets={},

    # Export per-gene codon usage tables by cluster
    export_cluster_codon_usage_enable=True,
    export_cluster_codon_usage_mode='RCU_devZ',
    export_cluster_codon_usage_codon_set='',
    export_cluster_codon_usage_include_genome_sheet=True,
    export_cluster_codon_usage_genome_sheet_name='Genome locus tags',
    export_cluster_codon_usage_round_decimals=6,
    export_cluster_codon_usage_raw_subdir='Raw codon usage tables',
    export_cluster_codon_usage_whole_genome_name='Whole genome',
    export_trna_usage_enable=False,
    trna_decoding_table_path='',
    trna_decoding_table_sheet='',
    export_trna_abundance_correlation_enable=False,
    trna_abundance_sheet='',
    trna_abundance_corr_show_fig=False,
    trna_abundance_corr_dpi=300,
    trna_abundance_heatmap_metric='ZTU',
    trna_abundance_scatter_metric='ZTU',
    trna_abundance_heatmap_clusters='all',
    trna_abundance_scatter_clusters='all',
    trna_abundance_scatter_yscale='linear',
    trna_abundance_scatter_show_fig=False,
    trna_gene_heatmap_enable=True,
    trna_gene_heatmap_metric='ZTU',
    trna_gene_heatmap_show_fig=True,
    trna_pairing_heatmap_enable=False,
    trna_pairing_heatmap_show_fig=False,
    trna_single_box_codon_heatmap_enable=True,
    trna_single_box_codon_heatmap_show_fig=True,
    trna_shift_heatmap_enable=True,
    trna_shift_heatmap_show_fig=True,
    trna_shift_heatmap_clusters='all',
    trna_wobble_heatmap_enable=True,
    trna_wobble_heatmap_show_fig=True,
    trna_wobble_heatmap_clusters='all',
    trna_modification_heatmap_enable=True,
    trna_modification_heatmap_show_fig=True,
    trna_mrna_stability_enable=True,
    trna_gene_wobble_plot_kind='heatmap',
    trna_gene_trna_plot_kind='heatmap',
    trna_mrna_stability_plot_kind='line',
    trna_gene_wobble_smooth=True,
    trna_gene_wobble_smooth_window=5,
    trna_gene_trna_smooth=True,
    trna_gene_trna_smooth_window=5,
    trna_mrna_stability_smooth=True,
    trna_mrna_stability_smooth_window=5,
    trna_wobble_plot_kind='boxplot',
    trna_shift_plot_kind='boxplot',
    trna_wobble_pair_stats_test='Student t-test',
    trna_shift_pair_stats_test='Student t-test',
    trna_wobble_pair_stats_gap=0.05,
    trna_shift_pair_stats_gap=0.05,
    trna_modifications_plot_kind='boxplot',
    trna_modifications_boxplot_caption_size=17,
    trna_secondary_axis_style='bars',
    trna_secondary_axis_alpha=0.22,
    trna_secondary_axis_bar_width=0.72,
    trna_boxplot_width=0.12,
    trna_modification_plots_box_width=0.18,
    trna_boxplot_show_points=True,
    trna_boxplot_point_alpha=0.35,
    trna_boxplot_point_size=10.5,
    trna_wobble_boxplot_style='boxplot',
    trna_shift_boxplot_style='boxplot',
    trna_modifications_boxplot_style='boxplot',
    trna_wobble_boxplot_log2=True,
    trna_shift_boxplot_log2=True,
    trna_modifications_boxplot_log2=True,
    trna_wobble_boxplot_ymin=None,
    trna_wobble_boxplot_ymax=None,
    trna_shift_boxplot_ymin=None,
    trna_shift_boxplot_ymax=None,
    trna_modifications_boxplot_ymin=None,
    trna_modifications_boxplot_ymax=None,
    trna_wobble_exclude_outliers=False,
    trna_shift_exclude_outliers=False,
    trna_modifications_exclude_outliers=False,
    trna_wobble_outlier_sd=3.0,
    trna_shift_outlier_sd=3.0,
    trna_modifications_outlier_sd=3.0,
    trna_modifications_feature_mode='modifications',
    trna_modifications_selected_features=None,   # None -> default feature set (modifications exclude ac4C34 only; m6A37 included); [] -> none
    trna_modifications_include_aas=None,         # None -> default Plot 6 amino-acid set: Ala/Arg/Asn/Asp/Cys/Gln/Glu/Gly/His/Ile/Leu/Lys/Phe/Pro/Ser/Thr/Tyr/Val; explicit list overrides
    trna_modifications_assignment_models='conservative,permissive',
    trna_modifications_stats_test='Student t-test',
    trna_modification_plots_caption_size=17,
    trna_modification_plots_dpi=300,
    trna_modification_plots_fig_width=18.0,
    trna_modification_plots_fig_height=8.0,
    trna_modification_plots_ymin=-2.1,
    trna_modification_plots_ymax=3.5,
    trna_modification_plots_group_bar_y=-0.26,
    trna_modification_plots_group_label_gap=0.07,
    trna_modification_plots_star_offset=0.07,
    trna_modification_plots_legend_ncol=None,
    trna_supp_heatmaps_customize=False,
    trna_supp_heatmaps_dpi=None,
    trna_supp_heatmaps_fig_width=None,
    trna_supp_heatmaps_fig_height=None,
    trna_supp_heatmaps_cell_height=None,
    trna_supp_heatmaps_xtick_every_genes=None,
    trna_supp_heatmaps_ytick_fontsize=None,
    trna_supp_heatmaps_title_fontsize=None,
    trna_supp_heatmaps_xmin=None,
    trna_supp_heatmaps_xmax=None,
    trna_supp_heatmaps_ymin=None,
    trna_supp_heatmaps_ymax=None,
    trna_shift_heatmaps_customize=False,
    trna_shift_heatmaps_dpi=None,
    trna_shift_heatmaps_fig_width=None,
    trna_shift_heatmaps_fig_height=None,
    trna_shift_heatmaps_cell_width=None,
    trna_shift_heatmaps_cell_height=None,
    trna_shift_heatmaps_xtick_fontsize=None,
    trna_shift_heatmaps_ytick_fontsize=None,
    trna_shift_heatmaps_title_fontsize=None,
    trna_shift_heatmaps_xmin=None,
    trna_shift_heatmaps_xmax=None,
    trna_shift_heatmaps_ymin=None,
    trna_shift_heatmaps_ymax=None,
    trna_shift_heatmap_log2_colorbar=True,
    trna_wobble_heatmap_log2_colorbar=True,
    trna_shift_heatmap_bracket_type='brace',
    trna_wobble_heatmap_bracket_type='brace',
    trna_shift_heatmap_bracket_x=-0.24,
    trna_shift_heatmap_label_x=-0.35,
    trna_wobble_heatmap_bracket_x=-0.17,
    trna_wobble_heatmap_label_x=-0.27,
    trna_shift_heatmap_bracket_lw=1.4,
    trna_wobble_heatmap_bracket_lw=1.4,
    plot_codon_gene_heatmap_enable=True,
    plot_codon_gene_heatmap_show_fig=True,
    figure_output_format='png',
    show_main_pipeline_figures=True,
    show_main_pipeline_scatter=False,
    ordered_show_mode=True,

    # Text outputs
    write_text_outputs=False,
)

RUNTIME_DEFAULTS = dict(
    usage_basis='RCU',
    codon_set='59',
    dimred_method='umap',
    cluster_method='kmeans',
    do_2d_ks=False,
    cluster_source='basic',
    basic_cluster_input_source='david_gene2terms',
    run_david_scan=False,
    compute_trna_usage=False,
    compute_trna_abundance_correlations=False,
    run_2d_density_plots=True,
    run_gchm=True,
)

BASIC_CLUSTER_KEYWORD_GROUPS = {
     "Ribosomal proteins": ["50S","30S","ribosome", "ribosomal", "ribonucleoprotein"],
     "Amino acid biosynthesis": ["biosynthesis of amino acids","amino acid biosynthesis", "amino-acid biosynthesis", "amino acid synthesis","biosynthesis of amino-acids", "amino-acid synthesis"],
     "Glycolysis - gluconeogenesis": ["glycolysis","glycolytic process","gluconeogenesis"],
     "TCA cycle": ["tca cycle", "tricarboxylic acid cycle", "citrate cycle", "citrate synthase", "aconitase", "isocitrate dehydrogenase", "α-ketoglutarate dehydrogenase", "succinyl-coa synthetase", "succinate dehydrogenase", "fumarase", "malate dehydrogenase", "oxidative decarboxylation"],
     "Carbon metabolism": ["ubiquinone", "tca", "carbon metabolism", "tricarboxylic acid", "citrate cycle","glycolysis"],
     "Cytochrome": ["cytochrome","cytochromes"],
     "Anaeroby": ["anaeroby", "anaerobic", "fumarate reductase", "nitrate reductase", "anaerobic respiration", "nitrite", "formate dehydrogenase", "pyruvate formate-lyase", "anaerobic growth"],
     "aa-tRNA synthetases":["synthetases","synthetase"],
     "tRNA charging": ["trnas", "trna", "aminoacyl-trna synthetase", "trna charging", "trna ligase", "trna synthetase", "aminoacylation"],
     "Cell Division": ["cell division", "cytokinesis", "septum", "fission", "binary fission", "divisome", "ftsz", "ftsa", "ftsi", "ftsq", "ftsk", "ftsl", "ftsn", "ftsw", "ftsb", "minc", "mind", "mine", "z-ring", "midcell", "division site"],
     "DNA mobility": ["transposition", "transposase", "integrase", "recombinase", "recombination"],
     "Transposase-Integrase":["transposase","tranpsos","IS110","integrase","DNA integration","transposition","IS1","insertions sequence"],
     "Restriction-Modification":["Restiction","Modification","restriction-modification"],
     "Strand exchange":["Resolvase","DNA strand exchange"],
     "Virulence": ["virulence","pathogen","pathogenicity", "pathogenesis", "toxin", "effector", "host", "infection", "invasion"],
     "Secretion": ["secretion", "secreted", "type iii", "type iv"],
     "Plasmids": ["plasmidic","plasmid","plasmids"],
     "Prophages": ["phage", "phages", "prophage", "prophages"],
     "LPS":["LPS","lipopolysaccharide","O-antigen"],
     "O-antigen": ["o antigen", "o-antigen"],
     "Motility": ["flagellum", "flagellar", "motility", "chemotaxis"],
     "Adhesion": ["adhesion", "fimbriae", "pili", "flagellum", "lps", "lipopolysaccharide"],
     "Quorum sensing": ["quorum sensing"],
     "Flagellum":["flagellum", "flagella"],
     "Chemotaxis":["chemotaxis","chemoreceptor","MCP","cheA","cheB","cheR","cheW","cheY","cheZ","cheC","cheD","cheV","cheX"],
     "helicases": ["helicase","ATP-dependent helicase","helicases"],
     "Leu-rich repeat (LRR) domains": ["LRR","Leucine-rich repeat","leu-rich","leu-rich repeat","leu-rich_rpt","LRR_dom","internalin","internalins"],
     "Nucleotide sugars biosynthesis":["Nucleotide sugars","amino sugars","nucleotide sugar","amino sugar"],
     "Major facilitator superfamily (MFS)":["MFS","major facilitator","solute binding","solute-binding"],
     "Secondary metabolites":["Secondary metabolites"],
     "Porins":["porins","porin"],
     "Inner Membrane": ["inner membrane", "inner membrane protein"],
     "Outer Membrane": ["outer membrane", "outer membrane protein"],
     "Acid stress": ["acid stress", "acid resistance", "low ph", "acid tolerance", "proton motive force", "glutamate decarboxylase", "gada", "gadb", "gadc", "acid shock"],
     "Heat shock": ["heat shock", "chaperone", "dnak", "groel", "groes", "grpe", "hsp", "heat inducible", "temperature stress", "thermotolerance"],
     "Oxidative stress": ["oxidative stress", "superoxide", "catalase", "soda", "sodb", "peroxidase", "hydrogen peroxide", "ros", "oxyr", "soxr", "soxs"],
     "Envelope stress": ["envelope stress", "cpx", "bacterial envelope", "degp", "sura", "sigmae", "omp", "misfolded protein", "outer membrane protein"],
     "Osmotic stress": ["osmotic stress", "osmoregulation", "osmoprotectant", "prop", "bett", "kdp", "osmy", "high salt", "nacl"],
     "Cold shock": ["cold shock", "cspa", "cspb", "low temperature", "cold inducible", "rna chaperone"],
     "Nutrient limitation": ["nutrient limitation", "nutrient starvation", "starvation response", "carbon starvation", "phosphate starvation", "phob", "phor", "camp", "crp", "nutrient stress"],
     "4Fe-4S": ["sulfur cluster", "4fe-4s"],
     "Aspartic peptidase AS" : ["Aspartic peptidase AS","Aspartic_peptidase","Aspartic_peptidase_AS"],
#    "virus/virions":["virus","viral","virion","virions"],
#    "Sporulation":["spore","sporulation","spore"],
#    "DNA binding proteins": ["dna-binding", "dna binding", "hth", "helix-turn-helix", "h-t-h", "~dna binding"],
#    "Transcriptional regulators": ["transcriptional regulator", "transcriptional factor", "transcriptional regulation"],
}

PIPELINE = dict(
    sheet_coordinates="Genes reordered",
    sheet_locus_tags="Locus Tags",
    sheet_binary="Binary",
    auto_run_plotting_pipeline=True,
    plotting_pipeline_script_path=os.path.join(HERE, "Plotting_Pipeline.py"),
    plot_include_columns=None,
    plot_highlight_columns=None,
    plot_max_nrows=2,
    codon_usage_plot_mode="Z",
    gchm_include_columns=None,
    per_cluster_suffix="_PerClusterGeneLists.xlsx",
)

KS_SETTINGS = dict(alpha=0.01, method="binned", bins=151, n_perm=2000, random_seed=42)


def _clean_string(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s


def _normalize_text(x):
    return " ".join(_clean_string(x).lower().split())


def _read_first_excel_sheet(path):
    return pd.read_excel(path, dtype=str)


def _clean_path_str(x):
    return str(x).strip() if x is not None else ""


def _move_replace(src, dst, retries=8, delay=0.35):
    src = _clean_path_str(src)
    dst = _clean_path_str(dst)
    if not src or not os.path.exists(src) or not dst:
        return ""
    if os.path.abspath(src) == os.path.abspath(dst):
        return dst

    os.makedirs(os.path.dirname(dst) or ".", exist_ok=True)

    last_err = None
    for attempt in range(int(max(1, retries))):
        try:
            if os.path.exists(dst):
                if os.path.isdir(dst):
                    shutil.rmtree(dst)
                else:
                    os.remove(dst)
            shutil.move(src, dst)
            return dst
        except PermissionError as e:
            last_err = e
            gc.collect()
            time.sleep(float(delay))
        except OSError as e:
            last_err = e
            gc.collect()
            time.sleep(float(delay))

    # Final best-effort fallback: copy then delete, which can sometimes succeed
    # on Windows after transient file locks are released.
    try:
        if os.path.exists(dst):
            if os.path.isdir(dst):
                shutil.rmtree(dst)
            else:
                os.remove(dst)
        shutil.copy2(src, dst)
        os.unlink(src)
        return dst
    except Exception:
        if last_err is not None:
            raise last_err
        raise


def _organize_pipeline_outputs(
    output_dir,
    artifacts,
    move_root_files=True,
    move_text_files=True,
    move_figure_files=True,
    move_david_outputs=True,
):
    output_dir = _clean_path_str(output_dir)
    if not output_dir or not os.path.isdir(output_dir):
        return {}

    # Figures and DAVID outputs are organized into dedicated subfolders.
    # Text/Methods exports are intentionally disabled for the public GUI workflow.
    # Do not create a Methods folder here.
    methods_dir = ""
    figures_dir = os.path.join(output_dir, "Figures")
    david_dir = os.path.join(output_dir, "DAVID sliding window analysis")
    for d in (figures_dir, david_dir):
        os.makedirs(d, exist_ok=True)

    moved = {}

    if move_text_files:
        text_paths = artifacts.get("text_paths") or {}
        if isinstance(text_paths, dict):
            text_map = {
                "methods_details": "Methods details.txt",
                "legends": "Legends.txt",
                "methods": "Methods.txt",
            }
            for key, new_name in text_map.items():
                src = text_paths.get(key, "")
                new_path = _move_replace(src, os.path.join(methods_dir, new_name))
                if new_path:
                    moved[key] = new_path
                    text_paths[key] = new_path

    if move_david_outputs:
        david_results = artifacts.get("david_results")

        if david_results:
            report_dir_src = david_results.get("report_dir", "")
            report_dir_dst = os.path.join(david_dir, "DAVID window reports")
            new_report_dir = _move_replace(report_dir_src, report_dir_dst)
            if new_report_dir:
                moved["report_dir"] = new_report_dir
                david_results["report_dir"] = new_report_dir

            david_moves = [
                (david_results.get("david_excel_path", ""), os.path.join(david_dir, "DAVID sliding window results.xlsx"), "david_excel_path"),
                (david_results.get("gene2terms_txt_path", ""), os.path.join(david_dir, "DAVID gene2terms.txt"), "gene2terms_txt_path"),
                (david_results.get("enrichment_plot_path", ""), os.path.join(david_dir, "DAVID enrichment scores.png"), "enrichment_plot_path"),
                (david_results.get("pvalue_plot_path", ""), os.path.join(david_dir, "Enrichment pvalues.png"), "pvalue_plot_path"),
            ]
            for src, dst, key in david_moves:
                new_path = _move_replace(src, dst)
                if new_path:
                    moved[key] = new_path
                    david_results[key] = new_path

    if move_root_files:
        root_renames = [
            (artifacts.get("main_workbook", ""), os.path.join(output_dir, "Clustering analysis results.xlsx"), "main_workbook"),
            (artifacts.get("per_cluster_workbook", ""), os.path.join(output_dir, "Gene lists per cluster.xlsx"), "per_cluster_workbook"),
            (artifacts.get("geneids_xlsx", ""), os.path.join(output_dir, "Gene IDs.xlsx"), "geneids_xlsx"),
            (artifacts.get("gc_txt_path", ""), os.path.join(output_dir, "GC content.txt"), "gc_txt_path"),
            (artifacts.get("cluster_codon_usage_xlsx", ""), os.path.join(output_dir, "Codon usage tables per cluster.xlsx"), "cluster_codon_usage_xlsx"),
        ]
        for src, dst, key in root_renames:
            new_path = _move_replace(src, dst)
            if new_path:
                moved[key] = new_path
                artifacts[key] = new_path

    if move_figure_files:
        for fname in list(os.listdir(output_dir)):
            src = os.path.join(output_dir, fname)
            if not os.path.isfile(src):
                continue
            figure_exts = ('.png', '.pdf', '.jpeg', '.jpg', '.tif', '.tiff')
            lower = fname.lower()
            if not lower.endswith(figure_exts):
                continue

            ext = os.path.splitext(fname)[1]
            if lower.startswith('gene_cluster_heatmap') and lower.endswith(figure_exts):
                new_name = 'Codons vs genes functional associations' + ext
            elif re.search(r'_main_density\.(png|pdf|jpe?g|tiff?)$', lower):
                new_name = '2D embedded scatter plot all genes' + ext
            elif '_local_' in lower:
                new_name = '2D embedded scatter plots per cluster' + ext
            elif re.search(r'_scatter\.(png|pdf|jpe?g|tiff?)$', lower):
                new_name = fname
            elif re.search(r'_heatmap\.(png|pdf|jpe?g|tiff?)$', lower):
                new_name = 'Codons vs genes 2D heatmap' + ext
            else:
                new_name = fname

            new_path = _move_replace(src, os.path.join(figures_dir, new_name))
            if new_path:
                moved.setdefault("figures", []).append(new_path)

        scatter_panel_dirs = []
        for fname in list(os.listdir(output_dir)):
            src_dir = os.path.join(output_dir, fname)
            if not os.path.isdir(src_dir):
                continue
            lower = fname.lower().strip()
            if lower.startswith('trna abundance vs ') and lower.endswith(' scatter panels'):
                scatter_panel_dirs.append(src_dir)

        for src_dir in scatter_panel_dirs:
            for fname in list(os.listdir(src_dir)):
                src = os.path.join(src_dir, fname)
                if not os.path.isfile(src):
                    continue
                new_path = _move_replace(src, os.path.join(figures_dir, fname))
                if new_path:
                    moved.setdefault("figures", []).append(new_path)
            try:
                if not os.listdir(src_dir):
                    os.rmdir(src_dir)
            except Exception:
                pass

    if moved:
        print("[INFO] Reorganized exported outputs:")
        for key, val in moved.items():
            if isinstance(val, list):
                print(f"  - {key}: {len(val)} file(s)")
            else:
                print(f"  - {key}: {val}")

    return moved


def _tk_dialogs_allowed() -> bool:
    return (Tk is not None) and (threading.current_thread() is threading.main_thread())

def _show_popup_message(title, message):
    try:
        if (not _tk_dialogs_allowed()) or messagebox is None:
            raise RuntimeError("tkinter messagebox unavailable")
        root = Tk()
        root.withdraw()
        messagebox.showinfo(title, message)
        root.destroy()
    except Exception:
        print(f"[INFO] {title}: {message}")


def _looks_like_david_unrecognized_ids_error(exc) -> bool:
    msg = str(exc or "").strip().lower()
    if msg == "":
        return False
    return (
        ("server raised fault" in msg and "index: 0, size: 0" in msg)
        or ("index: 0, size: 0" in msg)
        or ("no functional annotation" in msg)
        or ("unrecognized" in msg and "david" in msg)
        or ("non-annotated identifiers" in msg and "david" in msg)
        or ("locus tags and entrez gene ids" in msg and "david" in msg)
    )


def _show_david_unrecognized_ids_popup():
    _show_popup_message(
        "DAVID Bioinformatics",
        "The submitted genes do not appear to be recognized/annotated by DAVID Bioinformatics for this dataset.\n\n"
        "Note: codonpipe now tries DAVID first with locus tags, then falls back to Entrez Gene IDs when available.\n\n"
        "The DAVID sliding-window scan will be skipped. Please use the GeneIDs annotation source instead of DAVID gene2terms for basic clusters."
    )


def _choose_table_file(initialdir, title, filetypes):
    try:
        if (not _tk_dialogs_allowed()) or askopenfilename is None:
            raise RuntimeError("tkinter file dialog unavailable")
        root = Tk()
        root.withdraw()
        path = askopenfilename(
            title=title,
            initialdir=initialdir or None,
            filetypes=filetypes,
        )
        root.destroy()
    except Exception:
        path = input(f"{title}: ").strip().strip('"').strip("'")
    if not path:
        raise SystemExit("No file selected.")
    return path


def _choose_david_gene2terms_file(initialdir):
    return _choose_table_file(
        initialdir=initialdir,
        title="Select DAVID gene-to-terms TXT file",
        filetypes=[
            ("DAVID gene-to-terms TXT", "*.txt *.tsv *.csv"),
            ("Text files", "*.txt *.tsv *.csv"),
            ("All files", "*.*"),
        ],
    )


def _find_named_files(search_root, filename):
    matches = []
    if not search_root or not os.path.isdir(search_root):
        return matches
    for root, _, files in os.walk(search_root):
        for fname in files:
            if fname == filename:
                matches.append(os.path.join(root, fname))
    matches = sorted(set(matches), key=lambda p: os.path.getmtime(p), reverse=True)
    return matches


def _resolve_david_gene2terms_path(base_folder, output_dir):
    explicit_path = _clean_string(SET.get("david_gene2terms_path", ""))
    if explicit_path:
        explicit_path = os.path.abspath(os.path.expanduser(explicit_path))
        if os.path.isfile(explicit_path):
            print(f"[INFO] Using DAVID gene-to-terms TXT selected in GUI:\n  {explicit_path}")
            return explicit_path
        raise FileNotFoundError(f"DAVID gene-to-terms TXT file selected in GUI was not found: {explicit_path}")

    preferred_filename = str(SET.get("david_gene2terms_filename", "DAVID gene2terms.txt") or "DAVID gene2terms.txt")
    fallback_filenames = [preferred_filename]
    for alt in ["DAVID_gene2terms.txt", "DAVID gene2terms.txt"]:
        if alt not in fallback_filenames:
            fallback_filenames.append(alt)

    candidates = []
    for folder in [output_dir, base_folder]:
        for fname in fallback_filenames:
            for path in _find_named_files(folder, fname):
                if path not in candidates:
                    candidates.append(path)

    if not candidates:
        msg = (
            f"{preferred_filename} was not found. You must first run the DAVID sliding-window scan "
            f"to generate {preferred_filename}."
        )
        _show_popup_message("DAVID gene-to-terms file missing", msg)
        raise FileNotFoundError(msg)

    if len(candidates) == 1:
        print(f"[INFO] Using DAVID gene-to-terms TXT:\n  {candidates[0]}")
        return candidates[0]

    print("\n" + "=" * 70)
    print("Multiple DAVID gene-to-terms TXT files were found:")
    for i, path in enumerate(candidates, start=1):
        print(f"  {i}. {path}")
    print("  M. Choose another file manually")

    while True:
        ans = input("Enter file number (press Enter for 1): ").strip()
        if ans == "":
            return candidates[0]
        if ans.lower() == "m":
            return _choose_david_gene2terms_file(base_folder)
        if ans.isdigit():
            idx = int(ans) - 1
            if 0 <= idx < len(candidates):
                return candidates[idx]
        print("Invalid choice. Please enter a listed number or M.")


def _read_david_gene2terms_txt(path):
    df = pd.read_csv(path, sep="\t", dtype=str, keep_default_na=False)
    df = df.fillna("")
    required = {"DisplayLocusTag", "DAVID_Terms"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"DAVID gene-to-terms TXT is missing required columns: {sorted(missing)}"
        )
    return df

def _prompt_choice(title, options, default_key):
    key_to_value = {}
    number_to_value = {}
    default_number = None
    print("\n" + "=" * 70)
    print(title)
    for i, (key, label, desc, stored_value) in enumerate(options, start=1):
        default_tag = " [default]" if key == default_key else ""
        print(f"  {i}. {label}{default_tag}")
        if desc:
            print(f"     {desc}")
        key_to_value[str(key).strip().lower()] = stored_value
        key_to_value[str(label).strip().lower()] = stored_value
        key_to_value[str(stored_value).strip().lower()] = stored_value
        number_to_value[str(i)] = stored_value
        if key == default_key:
            default_number = str(i)

    if default_number is None:
        default_number = "1"

    while True:
        ans = input(f"Enter choice (press Enter for {default_number}): ").strip()
        if ans == "":
            return number_to_value[default_number]
        if ans in number_to_value:
            return number_to_value[ans]
        ans_l = ans.lower()
        if ans_l in key_to_value:
            return key_to_value[ans_l]
        print("Invalid choice. Please enter a listed option name or number.")


def _prompt_yes_no(title, description, default=False):
    default_number = "1" if default else "2"
    print("\n" + "=" * 70)
    print(title)
    if description:
        print(description)
    print("  1. YES")
    print(f"     {title}")
    action = str(title).strip().rstrip(" ?")
    print("  2. NO")
    print(f"     Do not {action[:1].lower() + action[1:]}")
    while True:
        ans = input(f"Enter choice (press Enter for {default_number}): ").strip().lower()
        if ans == "":
            return bool(default)
        if ans in {"1", "y", "yes", "true"}:
            return True
        if ans in {"2", "n", "no", "false", "0"}:
            return False
        print("Invalid choice. Please enter 1 or 2.")


def _looks_like_email(s):
    s = _clean_string(s)
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", s))


def _prompt_david_email_if_needed(current_email=""):
    current_email = _clean_string(current_email)

    if _looks_like_email(current_email):
        return current_email

    print("\n" + "=" * 70)
    print("DAVID Bioinformatics requires a registered email address to connect to the web service.")
    print("Please enter a valid DAVID-registered email address.")
    print("Press Enter without typing anything to skip the DAVID scan.")

    while True:
        ans = input("DAVID registered email: ").strip()
        if ans == "":
            return ""
        if _looks_like_email(ans):
            return ans
        print("This does not look like a valid email address. Please try again.")


def _choose_excel_sheet(path, purpose="Excel file"):
    xls = pd.ExcelFile(path)
    sheets = xls.sheet_names
    print("\n" + "=" * 70)
    print(f"Available sheets in {os.path.basename(path)} ({purpose}):")
    for i, s in enumerate(sheets, start=1):
        print(f"  {i}. {s}")
    ans = input("Enter sheet number to use (press Enter for first sheet): ").strip()
    if ans == "":
        return sheets[0]
    idx = int(ans) - 1
    if idx < 0 or idx >= len(sheets):
        raise ValueError("Invalid sheet selection.")
    return sheets[idx]


def _read_cluster_file_with_optional_sheet(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls', '.xlsm'):
        sheet = _choose_excel_sheet(path, purpose="refined cluster file")
        df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    else:
        df = pd.read_csv(path, sep=None, engine='python', dtype=str)
    return df.fillna("")


def _choose_output_id(row, preferred_col, fallback_cols):
    for col in [preferred_col] + list(fallback_cols):
        if col in row.index:
            val = _clean_string(row[col])
            if val and val not in {"NA", "None"}:
                return val
    return ""


def build_basic_clusters_from_annotation_df(annotation_df, keyword_groups, output_id_column,
                                            output_id_fallbacks, search_columns_preferred):
    df = annotation_df.copy().fillna("")
    search_columns = [c for c in search_columns_preferred if c in df.columns]
    if not search_columns:
        raise ValueError("None of the preferred search columns were found in the annotation file.")

    cluster_dict = {}
    long_rows = []
    for group_name, keywords in keyword_groups.items():
        matched_ids = []
        kw_norm = [_normalize_text(k) for k in keywords if _clean_string(k)]
        for _, row in df.iterrows():
            row_text = _normalize_text(" | ".join([_clean_string(row[c]) for c in search_columns]))
            matched_keywords = [kw for kw in kw_norm if kw in row_text]
            if not matched_keywords:
                continue
            output_id = _choose_output_id(row, output_id_column, output_id_fallbacks)
            if not output_id:
                continue
            matched_ids.append(output_id)
            long_rows.append({
                "Cluster": group_name,
                "MatchedID": output_id,
                "MatchedKeywords": "; ".join(matched_keywords),
                "GeneSymbol": _clean_string(row["GeneSymbol"]) if "GeneSymbol" in df.columns else "",
                "ProteinDescription": _clean_string(row["ProteinDescription"]) if "ProteinDescription" in df.columns else "",
                "EntrezGeneID": _clean_string(row["EntrezGeneID"]) if "EntrezGeneID" in df.columns else "",
                "RefSeq_LocusTag_RS": _clean_string(row["RefSeq_LocusTag_RS"]) if "RefSeq_LocusTag_RS" in df.columns else "",
                "LocusTag": _clean_string(row["LocusTag"]) if "LocusTag" in df.columns else "",
                "PrimaryID": _clean_string(row["PrimaryID"]) if "PrimaryID" in df.columns else "",
            })
        cluster_dict[group_name] = sorted(set(matched_ids))

    cluster_df = pd.DataFrame({group: pd.Series(ids) for group, ids in cluster_dict.items()})
    long_df = pd.DataFrame(long_rows)
    summary_df = pd.DataFrame([{"Cluster": g, "N_genes": len(ids)} for g, ids in cluster_dict.items()]).sort_values("N_genes", ascending=False)
    return cluster_df, long_df, summary_df


def _write_basic_cluster_workbook(cluster_df, long_df, summary_df, out_xlsx, annotation_source):
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        cluster_df.to_excel(writer, sheet_name="Clusters", index=False)
        long_df.to_excel(writer, sheet_name="Long_format_matches", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    print(f"[INFO] Automatic/basic clusters saved:\n  {out_xlsx}")
    print(f"[INFO] Automatic/basic cluster source:\n  {annotation_source}")
    return cluster_df, out_xlsx


def _current_basic_cluster_keyword_groups():
    """Return the active keyword dictionary for automatic cluster inference.

    The default is BASIC_CLUSTER_KEYWORD_GROUPS defined in this file. The GUI
    may override this at run time by setting SET["basic_cluster_keyword_groups"]
    to a plain {cluster_name: [keywords...]} dictionary. This avoids stale
    cluster definitions when users edit keywords in the interface.
    """
    groups = SET.get("basic_cluster_keyword_groups", None)
    if isinstance(groups, dict):
        cleaned = {}
        for name, keywords in groups.items():
            name_s = _clean_string(name)
            if not name_s:
                continue
            kws = [_clean_string(k) for k in list(keywords or []) if _clean_string(k)]
            if kws:
                cleaned[name_s] = kws
        # An explicitly empty dictionary means "no keyword-derived clusters".
        return cleaned
    return BASIC_CLUSTER_KEYWORD_GROUPS


def _build_basic_clusters_from_geneids_annotation(clustering_results, output_dir):
    geneids_xlsx = clustering_results.get("geneids_xlsx")
    geneids_df = clustering_results.get("geneids_df")
    if geneids_xlsx and os.path.isfile(geneids_xlsx):
        annotation_df = _read_first_excel_sheet(geneids_xlsx)
        annotation_source = geneids_xlsx
    elif geneids_df is not None and not geneids_df.empty:
        annotation_df = geneids_df.copy()
        annotation_source = "in-memory FASTA-derived GeneIDs table"
    else:
        raise ValueError("Automatic/basic cluster mode requested, but no GeneIDs annotation table is available.")

    keyword_groups = _current_basic_cluster_keyword_groups()
    if not keyword_groups:
        cluster_df = pd.DataFrame()
        long_df = pd.DataFrame()
        summary_df = pd.DataFrame()
    else:
        cluster_df, long_df, summary_df = build_basic_clusters_from_annotation_df(
            annotation_df=annotation_df,
            keyword_groups=keyword_groups,
            output_id_column=SET.get("basic_cluster_output_id_column", "RefSeq_LocusTag_RS"),
            output_id_fallbacks=SET.get("basic_cluster_output_id_fallbacks", ["GeneSymbol", "LocusTag", "PrimaryID"]),
            search_columns_preferred=SET.get("basic_cluster_search_columns", ["GeneSymbol", "ProteinDescription"]),
        )

    base_name = os.path.splitext(os.path.basename(clustering_results.get("fasta_path") or clustering_results.get("codon_file")))[0]
    out_xlsx = os.path.join(output_dir, f"{base_name}_basic_gene_clusters.xlsx")
    return _write_basic_cluster_workbook(cluster_df, long_df, summary_df, out_xlsx, annotation_source)


def _build_basic_clusters_from_david_gene2terms(clustering_results, output_dir, base_folder):
    david_txt_path = _resolve_david_gene2terms_path(base_folder, output_dir)
    annotation_df = _read_david_gene2terms_txt(david_txt_path)

    keyword_groups = _current_basic_cluster_keyword_groups()
    if not keyword_groups:
        cluster_df = pd.DataFrame()
        long_df = pd.DataFrame()
        summary_df = pd.DataFrame()
    else:
        cluster_df, long_df, summary_df = build_basic_clusters_from_annotation_df(
            annotation_df=annotation_df,
            keyword_groups=keyword_groups,
            output_id_column="DisplayLocusTag",
            output_id_fallbacks=["RefSeq_LocusTag_RS", "LocusTag", "PrimaryID", "GeneSymbol"],
            search_columns_preferred=["DAVID_Terms", "GeneSymbol", "ProteinDescription", "DisplayLocusTag"],
        )

    base_name = os.path.splitext(os.path.basename(clustering_results.get("fasta_path") or clustering_results.get("codon_file")))[0]
    out_xlsx = os.path.join(output_dir, f"{base_name}_basic_gene_clusters_from_DAVID_gene2terms.xlsx")
    return _write_basic_cluster_workbook(cluster_df, long_df, summary_df, out_xlsx, david_txt_path)


def _build_basic_clusters(clustering_results, output_dir, base_folder, input_source):
    src = str(input_source or "geneids").strip().lower()
    if src == "geneids":
        return _build_basic_clusters_from_geneids_annotation(clustering_results, output_dir)
    if src == "david_gene2terms":
        return _build_basic_clusters_from_david_gene2terms(clustering_results, output_dir, base_folder)
    raise ValueError(f"Unknown basic cluster input source: {input_source}")


def _cluster_sizes_from_df(cluster_df):
    sizes = {}
    for col in cluster_df.columns:
        vals = [
            v.strip() for v in cluster_df[col].replace({np.nan: ""}).astype(str).tolist()
            if v.strip() != "" and v.strip().lower() != "nan"
        ]
        sizes[str(col)] = len(dict.fromkeys(vals))
    return sizes


def _choose_clusters_for_display(cluster_df, min_genes_default):
    sizes = _cluster_sizes_from_df(cluster_df)
    ordered = list(cluster_df.columns)
    print("\n" + "=" * 70)
    print("Available clusters for display in the 2D enrichment plots and gene-cluster localization heatmap:")
    for i, col in enumerate(ordered, start=1):
        print(f"  {i}. {col} (n={sizes.get(str(col), 0)})")

    ans = input(
        "\nEnter cluster numbers or exact names separated by commas to display.\n"
        f"Press Enter to display all clusters with at least {int(min_genes_default)} genes: "
    ).strip()

    if ans == "":
        chosen = [str(col) for col in ordered if sizes.get(str(col), 0) >= int(min_genes_default)]
        return chosen if chosen else [str(col) for col in ordered]

    tokens = [t.strip() for t in re.split(r"[;,]+", ans) if t.strip()]
    lower_map = {str(col).strip().lower(): str(col) for col in ordered}
    chosen = []
    for tok in tokens:
        if tok.isdigit():
            idx = int(tok) - 1
            if 0 <= idx < len(ordered):
                chosen.append(str(ordered[idx]))
                continue
        mapped = lower_map.get(tok.lower())
        if mapped is not None:
            chosen.append(mapped)
        else:
            print(f"[WARN] Ignoring unknown cluster selection: {tok}")
    chosen = list(dict.fromkeys(chosen))
    return chosen if chosen else [str(col) for col in ordered]



def _suggest_plot_grid_rows(n_display_clusters: int) -> int:
    n_panels = 1 + max(0, int(n_display_clusters))
    return max(1, int(np.round(np.sqrt(n_panels))))


def _prompt_plot_grid_rows(default_rows=2):
    print("\n" + "=" * 70)
    print("Choose the number of rows for the 2D embedded plot showing all displayed clusters:")
    print("This controls the grid layout of the multi-panel figure with one panel per displayed cluster.")
    while True:
        ans = input(f"Enter number of rows (press Enter for {int(default_rows)}): ").strip()
        if ans == "":
            return max(1, int(default_rows))
        try:
            nrows = int(ans)
        except ValueError:
            print("Invalid choice. Please enter a positive integer.")
            continue
        if nrows < 1:
            print("Invalid choice. Please enter a positive integer.")
            continue
        return nrows


def _prompt_codon_usage_cluster_plot_mode(default_key="z"):
    print("\n" + "=" * 70)
    print("Choose whether to generate a multi-panel codon-usage plot per displayed cluster:")
    print("This produces horizontal bar plots with one panel per displayed cluster.")
    return _prompt_choice(
        "Select codon-usage plot type:",
        [
            ("ACU", "ACU", "Plot average absolute codon usage frequencies per panel.", "ACU"),
            ("RCU", "RCU", "Plot average relative codon usage frequencies within synonymous families.", "RCU"),
            ("Z",   "ZCU", "Plot average relative codon usage z-scores versus the whole genome.", "ZCU"),
            ("NONE", "No plot", "Do not generate the codon-usage multi-panel plot.", "NONE"),
        ],
        default_key=(default_key or "z").upper(),
    ).strip().upper()


def _prompt_organism_name(default_name=""):
    print("\n" + "=" * 70)
    print("Enter the name of the organism studied.")
    print("This will be used to create a workbook containing the displayed clusters as locus-tag columns.")
    ans = input(f"Organism name (press Enter for {default_name or 'Organism'}): ").strip()
    return ans if ans else (default_name or "Organism")


def _choose_trna_decoding_table(default_root=""):
    initialdir = str(default_root or SET.get("default_root", "") or os.getcwd())
    if askopenfilename is not None and _tk_dialogs_allowed():
        try:
            root = Tk() if Tk is not None else None
            if root is not None:
                root.withdraw()
                root.attributes("-topmost", True)
            path = askopenfilename(
                title="Select decoding-strategy Excel table",
                initialdir=initialdir if os.path.isdir(initialdir) else os.getcwd(),
                filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")],
            )
            if root is not None:
                root.destroy()
            if path:
                return path
        except Exception:
            pass
    ans = input("Path to codon-anticodon decoding Excel table (leave blank to cancel): ").strip()
    return ans


def _auto_find_trna_decoding_table(default_root="", fasta_path=""):
    """Find a likely unified decoding-strategy workbook in bundled/preloaded folders."""
    roots = []
    for value in [os.path.dirname(os.path.abspath(str(fasta_path or ""))) if fasta_path else "", default_root, SET.get("default_root", ""), ROOT, HERE]:
        value = str(value or "").strip()
        if value:
            roots.append(os.path.abspath(os.path.expanduser(value)))
    for sub in ["Preloaded genomes", "preloaded_genomes", "genomes"]:
        roots.append(os.path.join(HERE, sub))
        roots.append(os.path.join(ROOT, sub))

    seen = set()
    candidates = []
    for folder in roots:
        folder = os.path.abspath(os.path.expanduser(str(folder or "")))
        key = os.path.normcase(folder)
        if key in seen or not os.path.isdir(folder):
            continue
        seen.add(key)
        try:
            for dirpath, _dirnames, filenames in os.walk(folder):
                for fname in filenames:
                    low = fname.lower()
                    if not low.endswith((".xlsx", ".xls", ".xlsm")) or low.startswith("~$"):
                        continue
                    score_text = (fname + " " + os.path.basename(dirpath)).lower()
                    score = 0
                    for token, pts in [("trna", 6), ("codon", 5), ("anticodon", 5), ("decoding", 5), ("decode", 4), ("table", 1)]:
                        if token in score_text:
                            score += pts
                    if fasta_path and os.path.dirname(os.path.abspath(str(fasta_path))) == dirpath:
                        score += 3
                    if score >= 6:
                        candidates.append((-score, os.path.join(dirpath, fname)))
        except Exception:
            continue
    if not candidates:
        return ""
    candidates.sort(key=lambda item: (item[0], item[1].lower()))
    return candidates[0][1]


def _prompt_runtime_choices():
    usage_basis = _prompt_choice(
        "Choose which per-gene usage space to compute:",
        [
            ("RCU", "RCU", "Relative codon usage within synonymous codon families for each gene. Usually the best default for codon-bias structure.", "RCU"),
            ("ACU", "ACU", "Absolute codon usage / count-like codon composition per gene. Useful when raw codon abundance is the goal.", "ACU"),
            ("AA", "AA", "Amino-acid usage only. Collapses synonymous codons and removes codon-level information.", "AA"),
        ],
        default_key=RUNTIME_DEFAULTS["usage_basis"],
    )

    codon_set = {'AA': '64', 'ACU': '61', 'RCU': '59'}[str(usage_basis).upper()]
    print("\n" + "=" * 70)
    print("Codon set is selected automatically from usage basis to avoid mistakes.")
    print(f"Using codon set: {codon_set} (AA→64, ACU→61, RCU→59)")

    dimred_method = _prompt_choice(
        "Choose the dimensionality-reduction method:",
        [
            ("umap", "umap", "Non-linear embedding that usually gives the most useful visual separation for gene-level codon usage. Recommended default.", "umap"),
            ("tsne", "tsne", "Non-linear embedding that emphasizes local neighborhoods. Often visually strong, but less globally interpretable.", "tsne"),
            ("pca", "pca", "Linear dimensionality reduction. Fast and simple, with easier global interpretation.", "pca"),
        ],
        default_key=RUNTIME_DEFAULTS["dimred_method"],
    )

    cluster_method = _prompt_choice(
        "Choose the clustering method applied to the 2D embedding:",
        [
            ("kmeans", "kmeans", "Fast, stable partitioning into k groups. Recommended default.", "kmeans"),
            ("kmedoids", "kmedoids", "Medoid-based clustering, often a bit more robust to outliers, but may require sklearn-extra.", "kmedoids"),
            ("hierarchical", "hierarchical", "Hierarchical ordering of genes rather than a true multi-cluster partition in this pipeline.", "hierarchical"),
            ("dbscan", "dbscan", "Density-based clustering that can label noise points. Useful for irregular shapes.", "dbscan"),
            ("spectral", "spectral", "Graph-based clustering that can capture non-convex structures.", "spectral"),
        ],
        default_key=RUNTIME_DEFAULTS["cluster_method"],
    )

    do_2d_ks = _prompt_yes_no(
        "Compute the 2D Kolmogorov-Smirnov comparisons between displayed gene clusters?",
        "This compares whether selected clusters occupy significantly different regions of the 2D embedding. It adds extra statistics/tables but can take additional time.",
        default=RUNTIME_DEFAULTS["do_2d_ks"],
    )

    run_david_scan = _prompt_yes_no(
        "Run the DAVID sliding-window enrichment scan on the reordered genome?",
        "This scans the UMAP-reordered genome with overlapping windows, maps genes via the FASTA-derived GeneIDs table, and queries DAVID using locus tags first (LOCUS_TAG, then legacy symbol modes) before falling back to Entrez Gene IDs when needed.",
        default=RUNTIME_DEFAULTS["run_david_scan"],
    )

    compute_trna_usage = _prompt_yes_no(
        "Compute decoding-strategy tables in addition to codon-usage tables?",
        "This imports a unified decoding-strategy Excel table and computes per-gene decoding-group counts plus ATU, RTU, and ZTU values. The table must contain AA, Codon (5'-3'), and Anticodon (5'-3'); optional columns 4-7 provide tRNA abundance mean/STD, decoding modifications, and tRMEs.",
        default=RUNTIME_DEFAULTS.get("compute_trna_usage", False),
    )

    compute_trna_abundance_correlations = False
    if compute_trna_usage:
        compute_trna_abundance_correlations = _prompt_yes_no(
            "Compute tRNA-abundance correlation heatmaps when abundance columns are present?",
            "If the unified decoding table contains tRNA molecules/cell mean values in column 4, the pipeline will compute Pearson correlations between abundances and ATU / RTU / ZTU values, per amino acid and per cluster, and export heatmaps.",
            default=RUNTIME_DEFAULTS.get("compute_trna_abundance_correlations", False),
        )

    run_2d_density_plots = _prompt_yes_no(
        "Generate the cluster-specific 2D embedded panel figure?",
        "This creates the multi-panel figure with one embedded 2D panel per displayed cluster, plus the all-genes reference panel.",
        default=RUNTIME_DEFAULTS.get("run_2d_density_plots", True),
    )

    run_gchm = _prompt_yes_no(
        "Generate the gene-cluster localization heatmap along the genome axis?",
        "This creates the separate heatmap showing where the displayed clusters localize along the reordered genome axis.",
        default=RUNTIME_DEFAULTS.get("run_gchm", True),
    )

    cluster_source = _prompt_choice(
        "Choose how gene clusters should be defined:",
        [
            ("basic", "basic", "Automatically build broad/basic clusters from keyword matching on either the FASTA-derived GeneIDs table or a previously exported DAVID gene2terms.txt file.", "basic"),
            ("refined", "refined", "Load a refined cluster file from your computer via a file picker.", "refined"),
        ],
        default_key=RUNTIME_DEFAULTS["cluster_source"],
    )

    basic_cluster_input_source = None
    if cluster_source == "basic":
        basic_cluster_input_source = _prompt_choice(
            "Choose the annotation source used for keyword-based basic clusters:",
            [
                ("geneids", "geneids", "Use the FASTA-derived cds_from_genomic_geneIDs.xlsx-style annotation table, as in the current behavior.", "geneids"),
                ("david_gene2terms", "david_gene2terms", "Use a previously exported DAVID gene2terms.txt file.", "david_gene2terms"),
            ],
            default_key=RUNTIME_DEFAULTS["basic_cluster_input_source"],
        )

    organism_name = _prompt_organism_name(default_name="Organism")

    return dict(
        usage_basis=usage_basis,
        codon_set=codon_set,
        dimred_method=dimred_method,
        cluster_method=cluster_method,
        do_2d_ks=do_2d_ks,
        run_david_scan=run_david_scan,
        compute_trna_usage=compute_trna_usage,
        compute_trna_abundance_correlations=compute_trna_abundance_correlations,
        trna_abundance_heatmap_metric=trna_abundance_heatmap_metric,
        run_2d_density_plots=run_2d_density_plots,
        run_gchm=run_gchm,
        cluster_source=cluster_source,
        basic_cluster_input_source=basic_cluster_input_source,
        organism_name=organism_name,
    )


def _load_refined_cluster_df(base_folder):
    cluster_file_path = choose_cluster_file(base_folder)
    cluster_df = _read_cluster_file_with_optional_sheet(cluster_file_path)
    return cluster_df, cluster_file_path


def build_2dks_padj_matrix(ks_df: pd.DataFrame, cluster_order=None) -> pd.DataFrame:
    if ks_df is None or ks_df.empty:
        return pd.DataFrame()
    required = {"cluster_A", "cluster_B", "p_adj_BH"}
    missing = required - set(ks_df.columns)
    if missing:
        raise ValueError(f"2D KS raw table missing required columns: {sorted(missing)}")

    present = list(set(ks_df["cluster_A"].astype(str).tolist()) | set(ks_df["cluster_B"].astype(str).tolist()))
    present_set = set(present)
    if cluster_order is not None:
        ordered = [str(c) for c in list(cluster_order) if str(c) in present_set]
        names = ordered + [c for c in present if c not in set(ordered)]
    else:
        names = present
    if not names:
        return pd.DataFrame()

    mat = pd.DataFrame(np.ones((len(names), len(names))), index=names, columns=names, dtype=float)
    for _, r in ks_df.iterrows():
        a, b = str(r["cluster_A"]), str(r["cluster_B"])
        try:
            p = float(r["p_adj_BH"])
        except Exception:
            p = np.nan
        if a in mat.index and b in mat.columns:
            mat.loc[a, b] = p
            mat.loc[b, a] = p
    np.fill_diagonal(mat.values, 1.0)
    mat.index.name = 'cluster'
    return mat


def build_neglog10_matrix(p_mat: pd.DataFrame, min_p: float = 1e-300) -> pd.DataFrame:
    if p_mat is None or p_mat.empty:
        return pd.DataFrame()
    arr = p_mat.to_numpy(dtype=float)
    out = np.full_like(arr, np.nan, dtype=float)
    finite = np.isfinite(arr)
    out[finite] = -np.log10(np.clip(arr[finite], min_p, 1.0))
    neg = pd.DataFrame(out, index=p_mat.index, columns=p_mat.columns)
    neg.index.name = p_mat.index.name
    return neg

def _is_nonempty_id_value(v):
    if v is None:
        return False
    s = str(v).strip()
    return bool(s) and s.lower() not in {"nan", "na", "none"}


def _build_geneids_identifier_crosswalk(geneids_df):
    """Build a robust ID -> preferred locus-tag mapping from the FASTA-derived Gene IDs table.

    This is used to force all downstream sheets (coordinates, Binary, locus-tag sheets,
    metrics) into the same identifier namespace, even when the CDS FASTA row IDs were
    primary IDs rather than locus tags.
    """
    if geneids_df is None or geneids_df.empty:
        return {}

    df = geneids_df.copy()
    crosswalk = {}

    key_cols = [
        "LocusTag",
        "PrimaryID",
        "RefSeqProteinID",
        "RefSeq_LocusTag_RS",
        "Old_LocusTag",
    ]
    pref_cols = [
        "RefSeq_LocusTag_RS",
        "LocusTag",
        "Old_LocusTag",
        "PrimaryID",
        "RefSeqProteinID",
    ]

    for _, row in df.iterrows():
        preferred = ""
        for col in pref_cols:
            if col in df.columns and _is_nonempty_id_value(row.get(col, "")):
                preferred = str(row.get(col, "")).strip()
                break
        if not preferred:
            continue

        for col in key_cols:
            if col not in df.columns:
                continue
            raw = row.get(col, "")
            if not _is_nonempty_id_value(raw):
                continue
            key = str(raw).strip()
            crosswalk.setdefault(key, preferred)

    return crosswalk


def _canonicalize_ids_with_crosswalk(ids, alias_map, geneids_df=None):
    crosswalk = _build_geneids_identifier_crosswalk(geneids_df)
    out = []
    for raw in list(ids):
        raw_s = "" if raw is None else str(raw).strip()
        can_alias = canonicalize_id(raw_s, alias_map)

        preferred = ""
        if raw_s in crosswalk:
            preferred = crosswalk[raw_s]
        elif can_alias in crosswalk:
            preferred = crosswalk[can_alias]
        else:
            preferred = can_alias

        out.append(preferred if _is_nonempty_id_value(preferred) else raw_s)

    out = enforce_unique_after_canon(ids, out)
    return np.array(out, dtype=object)


def _run_david_scan_if_requested(runtime_choices, clustering_results, ordered_genes, alias_map, output_dir, strain_prefix):
    david_results = None
    geneids_df = clustering_results.get("geneids_df")

    if not runtime_choices.get("run_david_scan", False):
        return david_results

    if geneids_df is None or geneids_df.empty:
        print("[WARN] DAVID scan requested, but no FASTA-derived GeneIDs table is available. Skipping DAVID scan.")
        return david_results

    david_email = _prompt_david_email_if_needed(SET.get("david_user_email", ""))
    if david_email == "":
        print("[WARN] DAVID scan requested, but no valid DAVID email was provided. Skipping DAVID scan.")
        return david_results

    SET["david_user_email"] = david_email
    try:
        david_results = run_david_window_scan_from_ordered_genes(
            ordered_genes=list(ordered_genes),
            geneids_df=geneids_df,
            output_folder=output_dir,
            output_prefix=f"{strain_prefix}ClusteringAnalysis",
            user_email=david_email,
            alias_map=alias_map,
            geneids_xlsx_path=clustering_results.get("geneids_xlsx", None),
            window_size=int(SET.get("david_window_size", 100)),
            step_size=int(SET.get("david_step_size", 50)),
            wait_time=float(SET.get("david_wait_time", 0.0)),
            top_n_hits=int(SET.get("david_top_n_hits", 10)),
            max_clusters=int(SET.get("david_max_clusters", 3)),
            report_subdir_name=str(SET.get("david_report_subdir_name", "DAVID_window_reports")),
            min_valid_ids_per_window=int(SET.get("david_min_valid_ids_per_window", 3)),
            plot_format=str(SET.get("david_plot_format", "png")),
            chart_threshold=float(SET.get("david_chart_threshold", 1.0)),
            chart_count=int(SET.get("david_chart_count", 1)),
            manual_term_queries=list(SET.get("david_manual_term_queries", [])),
            term_match_mode=str(SET.get("david_term_match_mode", "contains")),
            append_to_geneids_excel=bool(SET.get("david_append_terms_to_geneids_excel", True)),
        )
    except Exception as e:
        print(f"[WARN] DAVID sliding-window scan failed: {e}")
        if _looks_like_david_unrecognized_ids_error(e):
            _show_david_unrecognized_ids_popup()
        david_results = None

    return david_results

def _sanitize_excel_sheet_name(name: str) -> str:
    name = str(name or "clusters").strip()
    name = re.sub(r"[\/*?:\[\]]", "_", name)
    return (name or "clusters")[:31]


def _next_cluster_sheet_name(existing_sheet_names) -> str:
    existing = {str(x) for x in (existing_sheet_names or [])}
    nums = []
    for s in existing:
        m = re.fullmatch(r"clusters-(\d+)", s.strip(), flags=re.IGNORECASE)
        if m:
            nums.append(int(m.group(1)))
    next_n = (max(nums) + 1) if nums else 1
    candidate = f"clusters-{next_n}"
    return _sanitize_excel_sheet_name(candidate)


def _next_unique_sheet_name(existing_sheet_names, preferred_name: str) -> str:
    existing = {str(x) for x in (existing_sheet_names or [])}
    pref = str(preferred_name or "Sheet").strip()
    pref_clean = _sanitize_excel_sheet_name(pref)
    if pref_clean.lower().startswith("clusters"):
        return _next_cluster_sheet_name(existing_sheet_names)
    if pref_clean not in existing:
        return pref_clean
    i = 2
    while True:
        suffix = f" {i}"
        base = pref_clean[: max(1, 31 - len(suffix))]
        cand = _sanitize_excel_sheet_name(base + suffix)
        if cand not in existing:
            return cand
        i += 1


def _nonempty_unique_values(values):
    seen = set()
    out = []
    for v in values:
        s = str(v).strip()
        if s == "" or s.lower() == "nan":
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def _filter_cluster_export_df(cluster_df, min_locus_tags=20):
    if cluster_df is None or getattr(cluster_df, "empty", True):
        return pd.DataFrame()
    export_cols = {}
    for col in cluster_df.columns:
        vals = _nonempty_unique_values(cluster_df[col].tolist())
        if len(vals) >= int(min_locus_tags):
            export_cols[str(col)] = pd.Series(vals, dtype=object)
    return pd.DataFrame(export_cols)


def _species_clusters_workbook_path(output_dir, organism_name):
    output_dir = _clean_path_str(output_dir)
    organism_name = _clean_string(organism_name) or "Organism"
    return os.path.join(output_dir, f"{organism_name} clusters.xlsx")


def _write_species_cluster_sheet(workbook_path, export_df, preferred_sheet_name):
    if export_df is None or getattr(export_df, "empty", True) or export_df.shape[1] == 0:
        return "", ""
    workbook_path = _clean_path_str(workbook_path)
    exists = os.path.isfile(workbook_path)
    existing_sheet_names = []
    if exists:
        try:
            wb = load_workbook(workbook_path)
            existing_sheet_names = list(wb.sheetnames)
            wb.close()
        except Exception:
            existing_sheet_names = []
    sheet_name = _next_unique_sheet_name(existing_sheet_names, preferred_sheet_name)
    mode = "a" if exists else "w"
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode=mode) as writer:
        export_df.to_excel(writer, sheet_name=sheet_name, index=False)
    return workbook_path, sheet_name


def _write_displayed_clusters_workbook(output_dir, organism_name, cluster_df, displayed_clusters, min_locus_tags=20):
    output_dir = _clean_path_str(output_dir)
    if not output_dir or cluster_df is None or cluster_df.empty:
        return "", ""

    chosen = [str(c) for c in (displayed_clusters or []) if str(c) in list(cluster_df.columns)]
    if not chosen:
        chosen = [str(c) for c in cluster_df.columns]

    export_df = cluster_df.loc[:, chosen].copy().fillna("")
    export_df = _filter_cluster_export_df(export_df, min_locus_tags=min_locus_tags)
    if export_df.empty or export_df.shape[1] == 0:
        print(f"[WARN] No displayed clusters reached the export threshold of {int(min_locus_tags)} locus tags; skipping species cluster workbook export.")
        return "", ""

    out_path = _species_clusters_workbook_path(output_dir, organism_name)
    out_path, sheet_name = _write_species_cluster_sheet(out_path, export_df, "clusters")
    if out_path:
        print(f"[INFO] Species-clusters workbook saved:\n  {out_path}\n  sheet: {sheet_name}")
    return out_path, sheet_name


def _append_david_derived_clusters_sheet(workbook_path, output_dir, organism_name, david_results, min_locus_tags=20):
    if not isinstance(david_results, dict):
        return workbook_path or "", ""
    term_cluster_df = david_results.get("term_cluster_df")
    if term_cluster_df is None or getattr(term_cluster_df, "empty", True):
        return workbook_path or "", ""

    export_df = _filter_cluster_export_df(term_cluster_df.copy().fillna(""), min_locus_tags=min_locus_tags)
    if export_df.empty or export_df.shape[1] == 0:
        print(f"[INFO] No DAVID-derived clusters reached the export threshold of {int(min_locus_tags)} locus tags; skipping DAVID derived clusters sheet.")
        return workbook_path or "", ""

    out_path = workbook_path or _species_clusters_workbook_path(output_dir, organism_name)
    out_path, sheet_name = _write_species_cluster_sheet(out_path, export_df, "DAVID derived clusters")
    if out_path:
        print(f"[INFO] DAVID-derived clusters appended to species workbook:\n  {out_path}\n  sheet: {sheet_name}")
    return out_path, sheet_name


# -----------------------------------------------------------------------------
# Optional custom CDS handling
# -----------------------------------------------------------------------------
_FASTA_EXTS_CUSTOM = (".fna", ".ffn", ".fa", ".fas", ".fasta")


def _split_custom_cds_paths(value):
    """Normalize GUI/SET custom-CDS path values to a de-duplicated list."""
    if value is None:
        return []
    raw = []
    if isinstance(value, (list, tuple, set)):
        raw = list(value)
    else:
        text = str(value).strip()
        if not text:
            return []
        for line in text.splitlines():
            raw.extend([x for x in line.split(";")])

    out, seen = [], set()
    for item in raw:
        path = str(item or "").strip().strip('"').strip("'")
        if not path:
            continue
        path = os.path.abspath(os.path.expanduser(path))
        key = os.path.normcase(path)
        if key in seen:
            continue
        seen.add(key)
        out.append(path)
    return out


def _iter_fasta_records_for_custom_merge(path):
    """Yield (header_without_gt, sequence) records from a FASTA file."""
    header = None
    seq_chunks = []
    with open(path, "r", encoding="utf-8", errors="ignore") as fh:
        for line in fh:
            line = line.rstrip("\n\r")
            if not line:
                continue
            if line.startswith(">"):
                if header is not None:
                    yield header, "".join(seq_chunks)
                header = line[1:].strip()
                seq_chunks = []
            else:
                seq_chunks.append(line.strip())
        if header is not None:
            yield header, "".join(seq_chunks)


def _clean_custom_cds_sequence(seq):
    s = re.sub(r"\s+", "", str(seq or "")).upper().replace("U", "T")
    return re.sub(r"[^ACGTN]", "N", s)


def _wrap_fasta_sequence(seq, width=80):
    return "\n".join(seq[i:i + width] for i in range(0, len(seq), width))


def _safe_header_text(text, max_len=180):
    s = re.sub(r"[\r\n\t]+", " ", str(text or "")).strip()
    s = s.replace("[", "(").replace("]", ")")
    s = " ".join(s.split())
    return s[:max_len]


def _safe_output_basename(text):
    s = str(text or "genome").strip()
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "genome"


def _prepare_custom_cds_fasta_if_requested(SET):
    """Append custom CDS FASTA records to the selected genome FASTA before analysis.

    Custom records are rewritten with stable, unique locus tags
    custom_CDS_001, custom_CDS_002, ... so that they survive all identifier
    canonicalization steps and can optionally be used as a cluster named
    ``custom``.
    """
    if not bool(SET.get("custom_cds_enable", False)):
        SET["custom_cds_generated_ids"] = []
        SET["custom_cds_original_fasta_path"] = ""
        SET["custom_cds_merged_fasta_path"] = ""
        return SET.get("fasta_path", "")

    base_fasta = str(SET.get("fasta_path", "") or SET.get("input_fasta", "")).strip()
    if not base_fasta or not os.path.isfile(base_fasta):
        raise FileNotFoundError(f"Base genome FASTA not found for custom-CDS merge: {base_fasta}")

    custom_paths = _split_custom_cds_paths(SET.get("custom_cds_paths", []))
    if not custom_paths:
        raise ValueError("'Add custom CDS to analysis' is enabled, but no custom FASTA/.fna file was selected.")

    for path in custom_paths:
        if not os.path.isfile(path):
            raise FileNotFoundError(f"Custom CDS FASTA file not found: {path}")
        if not path.lower().endswith(_FASTA_EXTS_CUSTOM):
            print(f"[WARN] Custom CDS file does not use a usual FASTA extension: {path}")

    custom_records = []
    counter = 0
    for path in custom_paths:
        n_in_file = 0
        for header, seq in _iter_fasta_records_for_custom_merge(path):
            clean_seq = _clean_custom_cds_sequence(seq)
            if not clean_seq:
                continue
            counter += 1
            n_in_file += 1
            custom_id = f"custom_CDS_{counter:03d}"
            product = _safe_header_text(f"Custom CDS from {os.path.basename(path)}; original header: {header}")
            custom_header = f">{custom_id} [locus_tag={custom_id}] [gene={custom_id}] [product={product}]"
            custom_records.append((custom_id, custom_header, clean_seq, path, header))
        if n_in_file == 0:
            print(f"[WARN] No readable CDS records found in custom FASTA file: {path}")

    if not custom_records:
        raise ValueError("No readable CDS records were found in the selected custom FASTA/.fna file(s).")

    output_root = str(SET.get("default_root", "") or "").strip()
    if not output_root:
        output_root = os.path.dirname(os.path.abspath(base_fasta))
    output_root = os.path.abspath(os.path.expanduser(output_root))
    merge_dir = os.path.join(output_root, "CodonPipe custom CDS merged FASTA")
    os.makedirs(merge_dir, exist_ok=True)

    base_name = _safe_output_basename(os.path.splitext(os.path.basename(base_fasta))[0])
    merged_path = os.path.join(merge_dir, f"{base_name}_with_custom_CDS.fna")

    with open(merged_path, "w", encoding="utf-8") as out:
        with open(base_fasta, "r", encoding="utf-8", errors="ignore") as base_fh:
            base_text = base_fh.read().rstrip()
            if base_text:
                out.write(base_text)
                out.write("\n")
        for custom_id, custom_header, clean_seq, _path, _orig_header in custom_records:
            out.write(custom_header + "\n")
            out.write(_wrap_fasta_sequence(clean_seq) + "\n")

    generated_ids = [r[0] for r in custom_records]
    SET["custom_cds_original_fasta_path"] = base_fasta
    SET["custom_cds_merged_fasta_path"] = merged_path
    SET["custom_cds_generated_ids"] = generated_ids
    SET["fasta_path"] = merged_path

    print("[INFO] Custom CDS merge enabled.")
    print(f"[INFO] Base genome FASTA: {base_fasta}")
    print(f"[INFO] Added {len(generated_ids)} custom CDS record(s) from {len(custom_paths)} file(s).")
    print(f"[INFO] Merged FASTA used for this run: {merged_path}")
    if bool(SET.get("custom_cds_include_as_cluster", False)):
        print(f"[INFO] Custom CDS will be included as cluster: {SET.get('custom_cds_cluster_name', 'custom')}")

    return merged_path


def _inject_custom_cds_cluster_if_requested(cluster_df, SET):
    """Add/extend the custom CDS cluster column in the cluster table."""
    if not bool(SET.get("custom_cds_enable", False)):
        return cluster_df
    if not bool(SET.get("custom_cds_include_as_cluster", False)):
        return cluster_df

    custom_ids = [str(x).strip() for x in list(SET.get("custom_cds_generated_ids", []) or []) if str(x).strip()]
    if not custom_ids:
        print("[WARN] Include-as-custom-cluster was requested, but no generated custom CDS IDs were available.")
        return cluster_df

    cluster_name = str(SET.get("custom_cds_cluster_name", "custom") or "custom").strip() or "custom"
    df = cluster_df.copy().fillna("") if cluster_df is not None else pd.DataFrame()

    existing_col = None
    for col in df.columns:
        if str(col).strip().lower() == cluster_name.lower():
            existing_col = col
            break

    existing_values = []
    if existing_col is not None:
        existing_values = [
            str(v).strip() for v in df[existing_col].replace({np.nan: ""}).astype(str).tolist()
            if str(v).strip() and str(v).strip().lower() != "nan"
        ]
        out_col = existing_col
    else:
        out_col = cluster_name

    merged_values = list(dict.fromkeys(existing_values + custom_ids))
    max_len = max(len(df), len(merged_values))
    df = df.reindex(range(max_len)).fillna("")
    df[out_col] = pd.Series(merged_values + [""] * (max_len - len(merged_values)), index=range(max_len))

    print(f"[INFO] Added custom CDS cluster '{out_col}' with {len(custom_ids)} gene(s).")
    return df



def _inject_fasta_metric_clusters_if_requested(cluster_df, SET, fasta_path, output_dir=None):
    """Append GUI-defined FASTA-derived metric clusters to the cluster table.

    The FASTA-derived clusters are intentionally not exported here, because this
    function runs before the final identifier canonicalization step. They are
    exported later with the same locus-tag namespace used by
    ``Gene lists per cluster.xlsx``.
    """
    configs = list(SET.get("fasta_metric_cluster_configs", []) or [])
    if (not configs) or (not bool(SET.get("fasta_metric_clusters_enable", False))):
        return cluster_df, pd.DataFrame(), pd.DataFrame()
    try:
        metric_cluster_df, scores_df = build_fasta_metric_cluster_df(
            fasta_path=fasta_path,
            metric_configs=configs,
            row_id_mode=str(SET.get("fasta_row_id_mode", "primary")),
            trim_to_multiple_of_3=bool(SET.get("fasta_trim_to_multiple_of_3", True)),
            organism_mode="prokaryote",
            codon_range=str(SET.get("fasta_codon_range", "all") or "all"),
        )
    except Exception as e:
        print(f"[WARN] Could not compute FASTA-derived metric clusters: {e}")
        return cluster_df, pd.DataFrame(), pd.DataFrame()

    if metric_cluster_df is None or metric_cluster_df.empty:
        print("[INFO] No FASTA-derived metric clusters were generated.")
        return cluster_df, pd.DataFrame(), scores_df if scores_df is not None else pd.DataFrame()

    print(f"[INFO] Added {metric_cluster_df.shape[1]} FASTA-derived metric cluster(s): {', '.join(map(str, metric_cluster_df.columns))}")
    return append_fasta_metric_clusters(cluster_df, metric_cluster_df), metric_cluster_df, (scores_df if scores_df is not None else pd.DataFrame())


def _sort_cluster_members_like_gene_lists(cluster_df, ordered_genes):
    """Sort cluster members exactly like the per-cluster gene-list workbook."""
    if cluster_df is None or getattr(cluster_df, "empty", True):
        return pd.DataFrame()

    ordered_iter = [] if ordered_genes is None else list(ordered_genes)
    ordered = [str(x).strip() for x in ordered_iter]
    order_index = {tag: i for i, tag in enumerate(ordered)}
    out_cols = {}
    max_len = 0
    for col in cluster_df.columns:
        vals = []
        seen = set()
        for raw in cluster_df[col].replace({np.nan: ""}).astype(str).tolist():
            v = str(raw).strip()
            if (not v) or v.lower() == "nan" or v in seen:
                continue
            seen.add(v)
            vals.append(v)
        vals.sort(key=lambda t: order_index.get(str(t), 10**12))
        out_cols[str(col)] = vals
        max_len = max(max_len, len(vals))

    out = pd.DataFrame(index=range(max_len)) if max_len else pd.DataFrame()
    for col, vals in out_cols.items():
        out[col] = pd.Series(vals + [""] * (max_len - len(vals)), index=range(max_len))
    return out.fillna("")


def _canonicalize_fasta_metric_outputs(metric_cluster_df, scores_df, alias_map, ordered_genes=None):
    """Return FASTA-derived metric groups/scores in final canonical locus-tag space.

    This mirrors the identifier handling used for the regular cluster sheets, so
    exported FASTA-derived groups contain the same locus-tag entries as the
    per-cluster gene-list workbook.
    """
    metric_can = pd.DataFrame()
    if metric_cluster_df is not None and not getattr(metric_cluster_df, "empty", True):
        metric_can = canonicalize_cluster_df(metric_cluster_df, alias_map)
        metric_can = _sort_cluster_members_like_gene_lists(metric_can, ordered_genes)

    scores_can = pd.DataFrame()
    if scores_df is not None and not getattr(scores_df, "empty", True):
        scores_can = scores_df.copy()
        if "LocusTag" in scores_can.columns:
            original = scores_can["LocusTag"].astype(str).map(lambda x: x.strip())
            canonical = original.map(lambda x: canonicalize_id(x, alias_map))
            if not original.equals(canonical):
                scores_can.insert(1, "OriginalLocusTag", original)
            scores_can["LocusTag"] = canonical
            if ordered_genes is not None:
                order_index = {str(tag): i for i, tag in enumerate(list(ordered_genes))}
                scores_can["__codonpipe_order"] = scores_can["LocusTag"].astype(str).map(lambda x: order_index.get(x, 10**12))
                scores_can = scores_can.sort_values(["__codonpipe_order", "LocusTag"], kind="stable").drop(columns=["__codonpipe_order"])
        scores_can = scores_can.reset_index(drop=True)

    return metric_can, scores_can


def _write_fasta_metric_outputs(output_dir, metric_cluster_df, scores_df, SET):
    """Export FASTA-derived metric clusters and per-gene scores to output_dir."""
    if (metric_cluster_df is None or getattr(metric_cluster_df, "empty", True)) and (scores_df is None or getattr(scores_df, "empty", True)):
        return "", ""

    out_dir = _clean_path_str(output_dir) or os.getcwd()
    os.makedirs(out_dir, exist_ok=True)
    cluster_path = os.path.join(out_dir, "FASTA-derived metric groups.xlsx")
    scores_path = os.path.join(out_dir, "FASTA-derived metric scores.xlsx")

    if metric_cluster_df is not None and not getattr(metric_cluster_df, "empty", True):
        with pd.ExcelWriter(cluster_path, engine="xlsxwriter") as writer:
            metric_cluster_df.to_excel(writer, sheet_name="Clusters", index=False)
    else:
        cluster_path = ""

    if scores_df is not None and not getattr(scores_df, "empty", True):
        with pd.ExcelWriter(scores_path, engine="xlsxwriter") as writer:
            scores_df.to_excel(writer, sheet_name="Metric scores", index=False)
            configs = list(SET.get("fasta_metric_cluster_configs", []) or [])
            if configs:
                pd.DataFrame(configs).to_excel(writer, sheet_name="Metric settings", index=False)
    else:
        scores_path = ""

    SET["fasta_metric_cluster_file_path"] = cluster_path
    SET["fasta_metric_cluster_scores_path"] = scores_path
    if cluster_path:
        print(f"[INFO] FASTA-derived metric clusters saved to:\n  {cluster_path}")
    if scores_path:
        print(f"[INFO] FASTA-derived metric scores saved to:\n  {scores_path}")
    return cluster_path, scores_path

def main():
    runtime_choices = _prompt_runtime_choices()
    SET["usage_basis"] = runtime_choices["usage_basis"]
    SET["codon_set"] = runtime_choices["codon_set"]
    SET["dimred_method"] = runtime_choices["dimred_method"]
    SET["cluster_method"] = runtime_choices["cluster_method"]
    SET["do_2d_ks"] = runtime_choices["do_2d_ks"]
    SET["organism_name"] = runtime_choices.get("organism_name", "Organism")

    # Optional GUI/backend feature: append user-supplied CDS FASTA records
    # to the selected genome FASTA before codon-usage tables are computed.
    _prepare_custom_cds_fasta_if_requested(SET)

    clustering_results = run_codon_clustering(SET)

    codon_file = clustering_results["codon_file"]
    prefix_hint = clustering_results.get("prefix_hint", "")
    gene_symbol_map = clustering_results.get("gene_symbol_map", {})
    gene_desc_map = clustering_results.get("gene_desc_map", {})
    geneids_df = clustering_results.get("geneids_df")

    row_names = clustering_results["RowNames"]
    Y = clustering_results["Y"]
    ordered_genes = clustering_results["ordered_genes"]
    AA_df = clustering_results["AA_df"]
    C_abs_df = clustering_results["C_abs_df"]
    C_rel_df = clustering_results["C_rel_df"]
    count_df = clustering_results.get("codon_count_df", C_abs_df)
    features_reorder = clustering_results["features_reorder"]
    output_dir = clustering_results["output_dir"]
    output_subfolder = clustering_results["output_subfolder"]
    strain_prefix = clustering_results["strain_prefix"]
    base_folder = os.path.dirname(codon_file)

    fasta_path = clustering_results.get("fasta_path")
    if fasta_path is None or (not os.path.isfile(str(fasta_path))):
        fasta_path = auto_find_fasta(base_folder, prefix_hint=prefix_hint)
        if fasta_path is None:
            fasta_path = choose_fasta(base_folder)

    locus_index, alias_map, id_map_df, missing_locus_headers, dup_counts = build_locus_index(
        fasta_path, organism_mode="prokaryote", codon_range=str(SET.get("fasta_codon_range", "all") or "all")
    )

    row_names = _canonicalize_ids_with_crosswalk(row_names, alias_map, geneids_df=geneids_df)
    ordered_genes = _canonicalize_ids_with_crosswalk(ordered_genes, alias_map, geneids_df=geneids_df)

    id_overlap = len(set(map(str, row_names.tolist())) & set(map(str, ordered_genes.tolist())))
    print(f"[INFO] Identifier overlap after canonicalization: {id_overlap}/{len(row_names)} coordinates matched to the reordered gene universe.")

    david_results = None
    needs_david_before_clusters = (
        runtime_choices.get("cluster_source") == "basic"
        and str(runtime_choices.get("basic_cluster_input_source", "geneids")).strip().lower() == "david_gene2terms"
        and runtime_choices.get("run_david_scan", False)
    )
    if needs_david_before_clusters:
        david_results = _run_david_scan_if_requested(
            runtime_choices=runtime_choices,
            clustering_results=clustering_results,
            ordered_genes=ordered_genes,
            alias_map=alias_map,
            output_dir=output_dir,
            strain_prefix=strain_prefix,
        )
        if david_results is None and str(runtime_choices.get("basic_cluster_input_source", "geneids")).strip().lower() == "david_gene2terms":
            print("[WARN] DAVID gene2terms could not be generated. Falling back to FASTA-derived GeneIDs annotations for basic clusters.")
            runtime_choices["basic_cluster_input_source"] = "geneids"

    if runtime_choices["cluster_source"] == "basic":
        cluster_df, cluster_file_path = _build_basic_clusters(
            clustering_results=clustering_results,
            output_dir=output_dir,
            base_folder=base_folder,
            input_source=runtime_choices.get("basic_cluster_input_source", "geneids"),
        )
    else:
        cluster_df, cluster_file_path = _load_refined_cluster_df(base_folder)

    cluster_df = _inject_custom_cds_cluster_if_requested(cluster_df, SET)
    cluster_df, fasta_metric_cluster_df_raw, fasta_metric_scores_df_raw = _inject_fasta_metric_clusters_if_requested(
        cluster_df, SET, fasta_path=fasta_path, output_dir=output_dir
    )

    plot_clusters = _choose_clusters_for_display(cluster_df, SET.get("plot_cluster_min_genes", 10))
    suggested_rows = _suggest_plot_grid_rows(len(plot_clusters))
    if bool(runtime_choices.get("run_2d_density_plots", True)):
        plot_max_nrows = _prompt_plot_grid_rows(default_rows=suggested_rows)
    else:
        plot_max_nrows = max(1, int(PIPELINE.get("plot_max_nrows", suggested_rows)))
    codon_usage_plot_mode = _prompt_codon_usage_cluster_plot_mode(default_key=PIPELINE.get("codon_usage_plot_mode", "Z"))
    PIPELINE["plot_include_columns"] = list(plot_clusters)
    PIPELINE["plot_highlight_columns"] = list(plot_clusters)
    PIPELINE["RUN_2D_DENSITY_PLOTS"] = bool(runtime_choices.get("run_2d_density_plots", True))
    SET["gchm_enable"] = bool(runtime_choices.get("run_gchm", SET.get("gchm_enable", True)))
    PIPELINE["auto_run_plotting_pipeline"] = bool(
        PIPELINE.get("RUN_2D_DENSITY_PLOTS", True)
        or SET.get("gchm_enable", True)
        or str(codon_usage_plot_mode).upper() != "NONE"
    )

    SET["export_trna_usage_enable"] = bool(runtime_choices.get("compute_trna_usage", False))
    if SET["export_trna_usage_enable"]:
        trna_path = str(SET.get("trna_decoding_table_path", "") or "").strip()
        if not trna_path:
            print("\n" + "=" * 70)
            print("Select the unified decoding-strategy Excel table used for decoding analyses.")
            print("Required columns: AA, Codon (5'-3'), Anticodon (5'-3').")
            print("Optional columns: tRNA molecules/cell mean, tRNA molecules/cell STD, tRNA modifications involved in decoding, tRMEs involved in decoding.")
            print("Pooled labels such as Leu UAA/CAA, Arg UCU/CCU, or Gly UCC/CCC are accepted.")
            print("CodonPipe will also try to find a matching Excel table in the 'Preloaded genomes' subfolder.")
            trna_path = _auto_find_trna_decoding_table(default_root=SET.get("default_root", ""), fasta_path=fasta_path)
            if trna_path:
                print(f"[INFO] Using preloaded decoding table:\n  {trna_path}")
            else:
                trna_path = _choose_trna_decoding_table(default_root=SET.get("default_root", ""))
        if not trna_path:
            raise RuntimeError("Decoding-strategy export was requested, but no unified decoding table was provided.")
        if not os.path.isfile(trna_path):
            raise FileNotFoundError(f"Decoding table not found: {trna_path}")
        SET["trna_decoding_table_path"] = trna_path
    else:
        SET["trna_decoding_table_path"] = ""
        SET["trna_decoding_table_sheet"] = ""
        SET["trna_abundance_sheet"] = ""
    displayed_clusters_workbook, displayed_clusters_sheet = _write_displayed_clusters_workbook(
        output_dir=output_dir,
        organism_name=runtime_choices.get("organism_name", "Organism"),
        cluster_df=cluster_df,
        displayed_clusters=plot_clusters,
        min_locus_tags=int(SET.get("plot_cluster_min_genes", 20)),
    )
    PIPELINE["plot_max_nrows"] = int(plot_max_nrows)
    PIPELINE["codon_usage_plot_mode"] = str(codon_usage_plot_mode).upper()
    PIPELINE["gchm_include_columns"] = list(plot_clusters)

    cluster_df = canonicalize_cluster_df(cluster_df, alias_map)
    gene_symbol_map = canonicalize_generic_map(gene_symbol_map, alias_map)
    gene_desc_map = canonicalize_generic_map(gene_desc_map, alias_map)

    fasta_metric_cluster_df, fasta_metric_scores_df = _canonicalize_fasta_metric_outputs(
        fasta_metric_cluster_df_raw, fasta_metric_scores_df_raw, alias_map=alias_map, ordered_genes=ordered_genes
    )
    fasta_metric_cluster_path, fasta_metric_scores_path = _write_fasta_metric_outputs(
        output_dir=output_dir,
        metric_cluster_df=fasta_metric_cluster_df,
        scores_df=fasta_metric_scores_df,
        SET=SET,
    )

    display_lower_map = {str(c).strip().lower(): str(c) for c in cluster_df.columns}
    display_clusters_can = []
    seen_display = set()
    for c in plot_clusters:
        mapped = display_lower_map.get(str(c).strip().lower())
        if mapped is not None and mapped not in seen_display:
            display_clusters_can.append(mapped)
            seen_display.add(mapped)
    if not display_clusters_can:
        display_clusters_can = list(cluster_df.columns)
    PIPELINE["plot_include_columns"] = list(display_clusters_can)
    PIPELINE["plot_highlight_columns"] = list(display_clusters_can)
    PIPELINE["gchm_include_columns"] = list(display_clusters_can)

    coords_df = build_coordinates_df(
        row_names, Y,
        ordered_genes=ordered_genes,
        gene_symbol_map=gene_symbol_map,
        gene_desc_map=gene_desc_map,
        locus_index=locus_index,
    )
    locus_tags_df = build_locus_tags_sheet(ordered_genes, cluster_df, genome_colname="Genome locus tags")
    binary_df = build_binary_sheet(ordered_genes, cluster_df, gene_symbol_map, locus_index)
    metrics_df, missing_list = build_metrics_table(ordered_genes, locus_index, gene_symbol_map)
    summary_df = build_summary(metrics_df, total_requested=len(ordered_genes), missing_count=len(missing_list))
    quantitative_bis_df = build_quantitative_bis(metrics_df)

    meta_df = build_meta_table(
        SET, PIPELINE, clustering_results,
        fasta_path=fasta_path,
        locustags_path=cluster_file_path,
        summary_df=summary_df,
        output_subfolder=output_subfolder
    )
    meta_extra = pd.DataFrame([
        {"Category": "Display", "Key": "Displayed clusters", "Value": ", ".join(display_clusters_can)},
        {"Category": "Display", "Key": "Cluster-specific 2D embedded panels", "Value": PIPELINE.get("RUN_2D_DENSITY_PLOTS", True)},
        {"Category": "Display", "Key": "2D embedded plot rows", "Value": PIPELINE.get("plot_max_nrows", 2)},
        {"Category": "Display", "Key": "Gene-cluster localization heatmap", "Value": SET.get("gchm_enable", True)},
        {"Category": "Display", "Key": "Codon-usage multi-panel plot", "Value": PIPELINE.get("codon_usage_plot_mode", "NONE")},
        {"Category": "Display", "Key": "Compute tRNA usage tables", "Value": SET.get("export_trna_usage_enable", False)},
        {"Category": "Display", "Key": "Compute tRNA-abundance correlations", "Value": SET.get("export_trna_abundance_correlation_enable", False)},
        {"Category": "Display", "Key": "Display threshold (min genes, when no manual selection)", "Value": SET.get("plot_cluster_min_genes", 10)},
        {"Category": "FASTA-derived metrics", "Key": "Metric groups workbook", "Value": fasta_metric_cluster_path},
        {"Category": "FASTA-derived metrics", "Key": "Metric scores workbook", "Value": fasta_metric_scores_path},
    ])
    meta_df = pd.concat([meta_df, meta_extra], ignore_index=True)

    ks_df = None
    if SET.get("do_2d_ks", False):
        ks_cluster_df = cluster_df.loc[:, [c for c in display_clusters_can if c in cluster_df.columns]] if display_clusters_can else cluster_df
        ks_df = compute_2d_ks_for_clusters(Y=Y, row_names=row_names, cluster_txt_df=ks_cluster_df, ks_settings=KS_SETTINGS)
    else:
        print("[INFO] 2D KS disabled by user choice; skipping.")

    if david_results is None:
        david_results = _run_david_scan_if_requested(
            runtime_choices=runtime_choices,
            clustering_results=clustering_results,
            ordered_genes=ordered_genes,
            alias_map=alias_map,
            output_dir=output_dir,
            strain_prefix=strain_prefix,
        )

    david_clusters_sheet = ""
    if david_results is not None:
        displayed_clusters_workbook, david_clusters_sheet = _append_david_derived_clusters_sheet(
            workbook_path=displayed_clusters_workbook,
            output_dir=output_dir,
            organism_name=runtime_choices.get("organism_name", "Organism"),
            david_results=david_results,
            min_locus_tags=int(SET.get("plot_cluster_min_genes", 20)),
        )
        meta_david = pd.DataFrame([
            {"Category": "DAVID", "Key": "Reports folder", "Value": david_results.get("report_dir", "")},
            {"Category": "DAVID", "Key": "Gene-to-terms TXT", "Value": david_results.get("gene2terms_txt_path", "")},
            {"Category": "DAVID", "Key": "Enrichment plot", "Value": david_results.get("enrichment_plot_path", "")},
            {"Category": "DAVID", "Key": "P-value plot", "Value": david_results.get("pvalue_plot_path", "")},
            {"Category": "DAVID", "Key": "Mapped genes", "Value": david_results.get("n_mapped_genes", "")},
            {"Category": "DAVID", "Key": "Total genes scanned", "Value": david_results.get("n_total_genes", "")},
            {"Category": "DAVID", "Key": "User email", "Value": SET.get("david_user_email", "")},
            {"Category": "DAVID", "Key": "Query terms used", "Value": "; ".join(david_results.get("query_terms", []))},
        ])
        meta_df = pd.concat([meta_df, meta_david], ignore_index=True)

    if displayed_clusters_workbook:
        meta_rows = [
            {"Category": "Display", "Key": "Displayed clusters workbook", "Value": displayed_clusters_workbook},
        ]
        if displayed_clusters_sheet:
            meta_rows.append({"Category": "Display", "Key": "Displayed clusters workbook sheet", "Value": displayed_clusters_sheet})
        if david_clusters_sheet:
            meta_rows.append({"Category": "Display", "Key": "DAVID derived clusters sheet", "Value": david_clusters_sheet})
        meta_df = pd.concat([
            meta_df,
            pd.DataFrame(meta_rows)
        ], ignore_index=True)

    out_base = os.path.join(output_dir, f"{strain_prefix}ClusteringAnalysis")
    out_xlsx = out_base + ".xlsx"

    AA_df_export = AA_df.copy()
    C_abs_df_export = C_abs_df.copy()
    C_rel_df_export = C_rel_df.copy()
    AA_df_export.index = row_names
    C_abs_df_export.index = row_names
    C_rel_df_export.index = row_names

    c_abs_idx_can = np.array([canonicalize_id(t, alias_map) for t in C_abs_df.index], dtype=object)
    c_abs_idx_can = enforce_unique_after_canon(C_abs_df.index, c_abs_idx_can)
    C_abs_df_can = C_abs_df.copy()
    C_abs_df_can.index = c_abs_idx_can
    C_rel_df_can = C_rel_df.copy()
    C_rel_df_can.index = c_abs_idx_can
    AA_df_can = AA_df.copy()
    AA_df_can.index = c_abs_idx_can

    count_idx_can = np.array([canonicalize_id(t, alias_map) for t in count_df.index], dtype=object)
    count_idx_can = enforce_unique_after_canon(count_df.index, count_idx_can)
    count_df_can = count_df.copy()
    count_df_can.index = count_idx_can

    with pd.ExcelWriter(out_xlsx, engine="xlsxwriter") as writer:
        coords_df.to_excel(writer, sheet_name=PIPELINE["sheet_coordinates"], index=False)

        feat_colname = "AA" if str(SET["usage_basis"]).upper() == "AA" else "Codon"
        pd.DataFrame({feat_colname: features_reorder}).to_excel(writer, sheet_name="Codons reordered", index=False)

        # Store the already computed codon/AA-vs-gene heatmap matrix so GUI figure
        # buttons can replot the heatmap later without rerunning dimensional
        # reduction/clustering or rewriting the main Excel workbook.
        heatmap_matrix = clustering_results.get("heatmap_matrix", None)
        if heatmap_matrix is not None:
            try:
                hm = pd.DataFrame(
                    np.asarray(heatmap_matrix, dtype=float),
                    index=[str(x) for x in list(features_reorder)],
                    columns=[str(x) for x in range(1, np.asarray(heatmap_matrix).shape[1] + 1)],
                )
                hm.insert(0, feat_colname, hm.index)
                hm.reset_index(drop=True).to_excel(writer, sheet_name="Heatmap matrix", index=False)
                pd.DataFrame([
                    {"Key": "title", "Value": clustering_results.get("heatmap_title", "Codons vs genes heatmap")},
                    {"Key": "usage_basis", "Value": SET.get("usage_basis", "")},
                    {"Key": "codon_set", "Value": SET.get("codon_set", "")},
                    {"Key": "dimred_method", "Value": SET.get("dimred_method", "")},
                    {"Key": "cluster_method", "Value": SET.get("cluster_method", "")},
                    {"Key": "bin_size", "Value": clustering_results.get("heatmap_bin_size", 1)},
                    {"Key": "n_genes_full", "Value": clustering_results.get("heatmap_n_genes_full", len(ordered_genes))},
                    {"Key": "matrix_rows", "Value": np.asarray(heatmap_matrix).shape[0]},
                    {"Key": "matrix_columns", "Value": np.asarray(heatmap_matrix).shape[1]},
                ]).to_excel(writer, sheet_name="Heatmap metadata", index=False)
            except Exception as e_hm_export:
                print(f"[WARN] Could not export heatmap replot matrix: {e_hm_export}")

        locus_tags_df.to_excel(writer, sheet_name=PIPELINE["sheet_locus_tags"], index=False)
        binary_df.to_excel(writer, sheet_name=PIPELINE["sheet_binary"], index=False)
        metrics_df.to_excel(writer, sheet_name="Quantitative", index=False)
        if quantitative_bis_df is None or quantitative_bis_df.empty:
            pd.DataFrame([{"info": "No binary (0/1) columns detected in 'Quantitative'."}]).to_excel(
                writer, sheet_name="Quantitative Locus tags", index=False
            )
        else:
            quantitative_bis_df.to_excel(writer, sheet_name="Quantitative Locus tags", index=False)
        if fasta_metric_cluster_df is not None and not fasta_metric_cluster_df.empty:
            fasta_metric_cluster_df.to_excel(writer, sheet_name="FASTA metric clusters", index=False)
        if fasta_metric_scores_df is not None and not fasta_metric_scores_df.empty:
            fasta_metric_scores_df.to_excel(writer, sheet_name="FASTA metric scores", index=False)
        meta_df.to_excel(writer, sheet_name="Meta", index=False)

        ks_raw_sheet = "2D KS - raw data"
        ks_cmp_sheet = "2D KS - comparison"
        if SET.get("do_2d_ks", False):
            if ks_df is not None and not ks_df.empty:
                ks_df.to_excel(writer, sheet_name=ks_raw_sheet, index=False)
                p_mat = build_2dks_padj_matrix(ks_df, cluster_order=cluster_df.columns)
                nlog_mat = build_neglog10_matrix(p_mat)
                pd.DataFrame([["Matrix of BH-adjusted p-values (p_adj_BH) for all cluster-pair comparisons (diagonal = 1)."]]).to_excel(
                    writer, sheet_name=ks_cmp_sheet, index=False, header=False, startrow=0, startcol=0
                )
                p_mat.to_excel(writer, sheet_name=ks_cmp_sheet, startrow=2, startcol=0, index=True)
                start_row_header2 = 2 + len(p_mat.index) + 3
                pd.DataFrame([["-log10(p_adj_BH) matrix (BH-adjusted p-values; diagonal = 0)."]]).to_excel(
                    writer, sheet_name=ks_cmp_sheet, index=False, header=False, startrow=start_row_header2, startcol=0
                )
                nlog_mat.to_excel(writer, sheet_name=ks_cmp_sheet, startrow=start_row_header2 + 2, startcol=0, index=True)
            else:
                msg = "2D KS analysis could not be computed (see console for details)."
                pd.DataFrame([[msg]]).to_excel(writer, sheet_name=ks_raw_sheet, index=False, header=False)
                pd.DataFrame([[msg]]).to_excel(writer, sheet_name=ks_cmp_sheet, index=False, header=False)

        if missing_list:
            pd.DataFrame({"missing_locus_tag": missing_list}).to_excel(writer, sheet_name="Missing", index=False)

        if david_results is not None:
            david_results['summary_df'].to_excel(writer, sheet_name='DAVID scan all', index=False)
            david_results['filtered_df'].to_excel(writer, sheet_name='DAVID filtered', index=False)
            david_results['top_hits_df'].to_excel(writer, sheet_name='top enrichment hits', index=False)
            david_results['mapping_df'].to_excel(writer, sheet_name='DAVID mapping', index=False)
            david_results['gene_term_df'].to_excel(writer, sheet_name='DAVID gene terms', index=False)
            david_results['genome_chart_df'].to_excel(writer, sheet_name='DAVID genome chart', index=False)
            david_results['auto_queries_df'].to_excel(writer, sheet_name='DAVID queries used', index=False)
            david_results['term_cluster_df'].to_excel(writer, sheet_name='DAVID term clusters', index=False)
            david_results['term_match_detail_df'].to_excel(writer, sheet_name='DAVID term matches', index=False)

    print(f"\n[INFO] Pipeline Excel saved:\n  {out_xlsx}")

    out_cluster_xlsx = out_base + PIPELINE.get("per_cluster_suffix", "_PerClusterGeneLists.xlsx")
    write_per_cluster_workbook(
        out_path=out_cluster_xlsx,
        cluster_df=cluster_df,
        ordered_genes=ordered_genes,
        gene_symbol_map=gene_symbol_map,
        gene_desc_map=gene_desc_map,
        row_names=row_names,
        Y=Y,
        dimred_method=SET.get("dimred_method", "umap"),
    )
    print(f"[INFO] Per-cluster Excel saved:\n  {out_cluster_xlsx}")

    out_usage = ""
    if bool(SET.get("export_cluster_codon_usage_enable", False)):
        try:
            out_usage = out_base + "__codon_usage_tables_per_cluster.xlsx"
            out_usage = write_codon_usage_by_cluster_workbook(
                out_path=out_usage,
                ordered_genes=list(ordered_genes),
                cluster_df=cluster_df,
                count_df=count_df_can,
                aa_df=AA_df_can,
                round_decimals=SET.get("export_cluster_codon_usage_round_decimals", 6),
                whole_genome_name=SET.get("export_cluster_codon_usage_whole_genome_name", "Whole genome"),
                raw_subdir_name=SET.get("export_cluster_codon_usage_raw_subdir", "Raw codon usage tables"),
                compute_trna_usage=bool(SET.get("export_trna_usage_enable", False)),
                trna_decoding_table_path=str(SET.get("trna_decoding_table_path", "") or ""),
                trna_decoding_table_sheet=SET.get("trna_decoding_table_sheet", ""),
            )
            print(f"[INFO] Codon-usage workbook saved:\n  {out_usage}")
            print(f"[INFO] Raw codon-usage workbooks saved in:\n  {os.path.join(output_dir, SET.get('export_cluster_codon_usage_raw_subdir', 'Raw codon usage tables'))}")
            if bool(SET.get("export_trna_usage_enable", False)):
                print(f"[INFO] tRNA-usage tables were also exported using:\n  {SET.get('trna_decoding_table_path', '')}")
                if bool(SET.get("export_trna_abundance_correlation_enable", False)):
                    try:
                        corr_outputs = write_trna_abundance_correlation_outputs(
                            summary_workbook_path=out_usage,
                            trna_decoding_table_path=str(SET.get("trna_decoding_table_path", "") or ""),
                            out_dir=output_dir,
                            trna_decoding_table_sheet=SET.get("trna_decoding_table_sheet", ""),
                            trna_abundance_sheet=SET.get("trna_abundance_sheet", ""),
                            write_workbook=False,
                            write_figures=True,
                            show_fig=bool(SET.get("trna_abundance_corr_show_fig", False)),
                            dpi=int(SET.get("trna_abundance_corr_dpi", SET.get("figure_dpi", 300))),
                            heatmap_metric=str(SET.get("trna_abundance_heatmap_metric", "ZTU") or "ZTU"),
                            scatter_metric=str(SET.get("trna_abundance_scatter_metric", "ZTU") or "ZTU"),
                            scatter_yscale=str(SET.get("trna_abundance_scatter_yscale", "linear") or "linear"),
                            scatter_show_fig=bool(SET.get("trna_abundance_scatter_show_fig", False)),
                            heatmap_clusters=str(SET.get("trna_abundance_heatmap_clusters", "all") or "all"),
                            scatter_clusters=str(SET.get("trna_abundance_scatter_clusters", "all") or "all"),
                            heatmap_xmin=SET.get("trna_abundance_heatmap_xmin", None),
                            heatmap_xmax=SET.get("trna_abundance_heatmap_xmax", None),
                            heatmap_ymin=SET.get("trna_abundance_heatmap_ymin", None),
                            heatmap_ymax=SET.get("trna_abundance_heatmap_ymax", None),
                            scatter_xmin=SET.get("trna_abundance_scatter_xmin", None),
                            scatter_xmax=SET.get("trna_abundance_scatter_xmax", None),
                            scatter_ymin=SET.get("trna_abundance_scatter_ymin", None),
                            scatter_ymax=SET.get("trna_abundance_scatter_ymax", None),
                        )
                        if corr_outputs:
                            if corr_outputs.get("workbook"):
                                print(f"[INFO] tRNA-abundance correlation workbook saved:\n  {corr_outputs.get('workbook', '')}")
                            heatmap_paths = [v for k, v in corr_outputs.items() if str(k).startswith('figure_')]
                            scatter_paths = [v for k, v in corr_outputs.items() if str(k).startswith('scatter_')]
                            if heatmap_paths or scatter_paths:
                                print(f"[INFO] tRNA-abundance correlation figures saved: {len(heatmap_paths)} heatmap(s), {len(scatter_paths)} scatter-panel figure(s)")
                        else:
                            print("[INFO] No tRNA abundance values were detected in column 4 of the decoding table; skipping tRNA-abundance correlation outputs.")
                    except Exception as e_corr:
                        print(f"[WARN] Could not export tRNA-abundance correlation outputs: {e_corr}")
                try:
                    trna_stage1_set = dict(SET)
                    trna_stage1_set['trna_shift_heatmap_enable'] = False
                    trna_stage1_set['trna_wobble_heatmap_enable'] = False
                    trna_stage1_set['trna_modification_heatmap_enable'] = False
                    trna_stage1_outputs = render_trna_gene_ordered_heatmaps(
                        SET=trna_stage1_set,
                        count_df=count_df_can,
                        ordered_genes=list(ordered_genes),
                        output_dir=output_dir,
                        cluster_df=cluster_df,
                        selected_clusters=list(display_clusters_can),
                    )
                    if trna_stage1_outputs:
                        print(f"[INFO] Gene-ordered tRNA/codon heatmaps saved: {len(trna_stage1_outputs)} figure(s)")
                except Exception as e_trna_fig:
                    print(f"[WARN] Could not export gene-ordered tRNA heatmaps: {e_trna_fig}")
        except Exception as e:
            print(f"[WARN] Could not export codon-usage workbook: {e}")

    text_paths = {}
    if False and bool(SET.get("write_text_outputs", False)):
        try:
            paths = write_all_texts(
                out_base=out_base,
                SET=SET,
                PIPELINE=PIPELINE,
                KS_SETTINGS=KS_SETTINGS,
                clustering_results=clustering_results,
                codon_file=codon_file,
                cluster_file_path=cluster_file_path,
                fasta_path=fasta_path,
                out_xlsx=out_xlsx,
                out_cluster_xlsx=out_cluster_xlsx,
                out_codon_usage_xlsx=out_usage,
            )
            text_paths = dict(paths)
            print("[INFO] Text outputs saved:")
            for k, p in paths.items():
                print(f"  - {k}: {p}")
        except Exception as e:
            print(f"[WARN] Could not write text outputs: {e}")

    artifacts_for_organization = dict(
        main_workbook=out_xlsx,
        per_cluster_workbook=out_cluster_xlsx,
        codon_xlsx="",
        geneids_xlsx=clustering_results.get("geneids_xlsx", ""),
        gc_txt_path=clustering_results.get("gc_txt_path", ""),
        cluster_codon_usage_xlsx=out_usage,
        displayed_clusters_workbook=displayed_clusters_workbook,
        text_paths=text_paths,
        david_results=david_results,
        cluster_file_path=cluster_file_path,
    )

    _organize_pipeline_outputs(
        output_dir=output_dir,
        artifacts=artifacts_for_organization,
        move_root_files=True,
        move_text_files=False,
        move_figure_files=False,
        move_david_outputs=True,
    )
    out_xlsx_final = artifacts_for_organization.get("main_workbook", out_xlsx) or out_xlsx

    if PIPELINE.get("auto_run_plotting_pipeline", False):
        print(f"[INFO] Auto-run plotting enabled. Using script:\n  {PIPELINE.get('plotting_pipeline_script_path', '')}")
        old_ob = os.environ.get("CODONPIPE_OUT_BASE")
        old_dm = os.environ.get("CODONPIPE_DIMRED_METHOD")
        _codon_plot_keys = [
            "CODONPIPE_CODON_USAGE_PLOT_MODE",
            "CODONPIPE_CODON_USAGE_WORKBOOK",
        ]
        old_codon_plot = {k: os.environ.get(k) for k in _codon_plot_keys}
        _gchm_keys = [
            "CODONPIPE_GCHM_ENABLE", "CODONPIPE_GCHM_SHEET", "CODONPIPE_GCHM_COLORMAP",
            "CODONPIPE_GCHM_CUSTOM_CMAPS_XLSX", "CODONPIPE_GCHM_SIGMA", "CODONPIPE_GCHM_SPREAD_FACTOR",
            "CODONPIPE_GCHM_HEIGHT_PER_CLUSTER", "CODONPIPE_GCHM_LABEL_FONTSIZE", "CODONPIPE_GCHM_DPI",
            "CODONPIPE_GCHM_CMAP_MIN_REL", "CODONPIPE_GCHM_CMAP_MAX_REL", "CODONPIPE_GCHM_OUTPUT_FILENAME",
            "CODONPIPE_GCHM_SHOW",
        ]
        old_gchm = {k: os.environ.get(k) for k in _gchm_keys}
        os.environ["CODONPIPE_OUT_BASE"] = out_base
        os.environ["CODONPIPE_DIMRED_METHOD"] = str(SET.get("dimred_method", "umap"))
        os.environ["CODONPIPE_CODON_USAGE_PLOT_MODE"] = str(PIPELINE.get("codon_usage_plot_mode", "NONE") or "NONE")
        os.environ["CODONPIPE_CODON_USAGE_WORKBOOK"] = str(artifacts_for_organization.get("cluster_codon_usage_xlsx", out_usage) or "")
        os.environ["CODONPIPE_GCHM_ENABLE"] = "1" if bool(SET.get("gchm_enable", True)) else "0"
        os.environ["CODONPIPE_GCHM_SHEET"] = str(PIPELINE.get("sheet_locus_tags", "Locus Tags"))
        os.environ["CODONPIPE_GCHM_COLORMAP"] = str(SET.get("gchm_colormap", "plasma"))
        os.environ["CODONPIPE_GCHM_CUSTOM_CMAPS_XLSX"] = str(SET.get("gchm_custom_cmaps_xlsx", "") or "")
        os.environ["CODONPIPE_GCHM_SIGMA"] = str(SET.get("gchm_sigma", 10))
        os.environ["CODONPIPE_GCHM_SPREAD_FACTOR"] = str(SET.get("gchm_spread_factor", 5))
        os.environ["CODONPIPE_GCHM_HEIGHT_PER_CLUSTER"] = str(SET.get("gchm_height_per_cluster", 0.3))
        os.environ["CODONPIPE_GCHM_LABEL_FONTSIZE"] = str(SET.get("gchm_label_fontsize", 10))
        os.environ["CODONPIPE_GCHM_DPI"] = str(SET.get("gchm_dpi", 300))
        os.environ["CODONPIPE_GCHM_CMAP_MIN_REL"] = str(SET.get("gchm_cmap_min_rel", 0.2))
        os.environ["CODONPIPE_GCHM_CMAP_MAX_REL"] = str(SET.get("gchm_cmap_max_rel", 1.0))
        os.environ["CODONPIPE_GCHM_OUTPUT_FILENAME"] = str(SET.get("gchm_output_filename", "gene_cluster_heatmap_KS.png"))
        os.environ["CODONPIPE_GCHM_SHOW"] = "1" if (bool(SET.get("gchm_show_fig", True)) and (not bool(SET.get("ordered_show_mode", True)))) else "0"
        try:
            run_density_plot_script(out_xlsx_final, PIPELINE)
        finally:
            if old_ob is None:
                os.environ.pop("CODONPIPE_OUT_BASE", None)
            else:
                os.environ["CODONPIPE_OUT_BASE"] = old_ob
            if old_dm is None:
                os.environ.pop("CODONPIPE_DIMRED_METHOD", None)
            else:
                os.environ["CODONPIPE_DIMRED_METHOD"] = old_dm
            for k, v in old_codon_plot.items():
                if v is None:
                    os.environ.pop(k, None)
                else:
                    os.environ[k] = v
            for k, v in old_gchm.items():
                if v is None:
                    os.environ.pop(k, None)
                else:
                    os.environ[k] = v


    if bool(SET.get('export_trna_usage_enable', False)):
        try:
            trna_stage2_set = dict(SET)
            trna_stage2_set['trna_gene_heatmap_enable'] = False
            trna_stage2_set['trna_single_box_codon_heatmap_enable'] = False
            trna_stage2_set['trna_shift_heatmap_enable'] = bool(SET.get('trna_shift_heatmap_enable', True))
            trna_stage2_set['trna_wobble_heatmap_enable'] = bool(SET.get('trna_wobble_heatmap_enable', True))
            trna_stage2_set['trna_modification_heatmap_enable'] = bool(SET.get('trna_modification_heatmap_enable', True))
            trna_stage2_outputs = render_trna_gene_ordered_heatmaps(
                SET=trna_stage2_set,
                count_df=count_df_can,
                ordered_genes=list(ordered_genes),
                output_dir=output_dir,
                cluster_df=cluster_df,
                selected_clusters=list(display_clusters_can),
            )
            stage2_only = {k: v for k, v in (trna_stage2_outputs or {}).items() if k in {'trna_usage_cluster_shift', 'wobble_cluster_enrichment', 'trna_modification_cluster_shift', 'trna_modification_cluster_shift_conservative', 'trna_modification_cluster_shift_permissive', 'trna_modification_cluster_shift_legacy', 'trna_shift_heatmap', 'trna_wobble_shift_heatmap', 'trna_modification_heatmap'}}
            if stage2_only:
                print(f"[INFO] Cluster-level decoding enrichment plots saved: {len(stage2_only)} figure(s)")
        except Exception as e_trna_stage2:
            print(f"[WARN] Could not export cluster-level tRNA/codon enrichment heatmaps: {e_trna_stage2}")

    _organize_pipeline_outputs(
        output_dir=output_dir,
        artifacts=artifacts_for_organization,
        move_root_files=False,
        move_text_files=False,
        move_figure_files=True,
        move_david_outputs=False,
    )

    print("\n[DONE] Full codon-usage → clusters → quantitative metrics pipeline finished.")
    plt.show()


if __name__ == "__main__":
    main()